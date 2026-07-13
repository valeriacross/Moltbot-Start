import os, io, logging, telebot, html, time, random, threading
from PIL import Image
from telebot import types
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from google.genai import types as genai_types
from C_shared100 import GeminiClient, CaptionGenerator, HealthServer, is_allowed, SHARED_VERSION, SHARED_DATE
from C_shared100 import VALERIA_FACE, VALERIA_BODY_STRONG, VALERIA_BODY_SAFE, VALERIA_WATERMARK, VALERIA_NEGATIVE
from C_shared100 import VALERIA_DNA, build_valeria_identity, analyze_scene

# --- VERSIONE ---
# CHANGELOG 2.0.1 (13/07/2026): rimosse MODEL_ID e MODEL_TEXT_ID — dead
# code, verificato via grep su tutto il file: mai referenziate fuori dalla
# propria dichiarazione, /info di Filtro non mostra alcuna stringa modello
# (a differenza di Vogue/Atelier/Surprise). Nessuna modifica funzionale —
# Filtro chiama sempre e solo GeminiClient.generate()/analyze_scene(), che
# usano MODEL definito in C_shared100.py, mai queste due costanti locali.
# CHANGELOG 2.0.0 (20/06/2026): fix #9 — rimosso blocco cleanup mosaic
# irraggiungibile in handle_photo (dead code, verificato non necessario).
VERSION = "2.0.1"

# --- LOGGING ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- CONFIGURAZIONE ---
TOKEN        = os.environ.get("TELEGRAM_TOKEN_FX")

gemini  = GeminiClient()
caption = CaptionGenerator(gemini)
server  = HealthServer("Filtro", VERSION)

# --- STATO UTENTE ---
user_filter   = defaultdict(lambda: None)   # filtro selezionato
user_category = defaultdict(lambda: None)   # categoria selezionata
user_artist   = {}                          # {uid: str} — artista selezionato per artistic_style
lego_waiting_list = {}                      # {uid: {'counts': dict, 'cols': int, 'rows': int}}
pending       = {}                          # {uid: {'img': bytes, 'filter_key': str}}
last_img      = {}                          # {uid: bytes} — ultima immagine usata
last_prompt   = {}                          # {uid: str} — ultimo prompt inviato all'API

executor = ThreadPoolExecutor(max_workers=4)

# --- STATO RACCOLTA MOSAIC ---
mosaic_collecting = {}   # {uid: {'photos': [bytes,...], 'timer': Timer}}

# --- STILI ARTISTICI (per filtro artistic_style) ---
ARTISTIC_STYLES_POOL = [
    # ── RINASCIMENTO & CLASSICI ─────────────────────────────────
    "🖼️ Leonardo — sfumato technique, atmospheric depth, Renaissance masterwork",
    "🕊️ Raffaello — harmonious Renaissance composition, soft idealized grace, warm palette",
    "⚡ Michelangelo — monumental sculptural figures, dramatic musculature, Sistine grandeur",
    "🕯️ Caravaggio — extreme chiaroscuro, tenebrism, dramatic realism, Baroque contrast",
    # ── IMPRESSIONISMO & POST ───────────────────────────────────
    "🌸 Renoir — soft impressionist light, warm skin tones, dappled outdoor atmosphere",
    "🌻 Van Gogh — swirling expressive brushstrokes, vivid saturated palette, emotional intensity",
    "🐟 Matisse — bold flat color fields, joyful decorative patterns, Fauvist energy",
    "🌙 Chagall — dreamlike floating figures, folkloric symbolism, luminous jewel tones",
    # ── MODERNISMO & ASTRATTO ───────────────────────────────────
    "✨ Klimt — gold leaf ornamental surfaces, Byzantine decorative excess, Viennese Secession",
    "⭕ Mirò — playful biomorphic shapes, primary colors on white, Catalan surrealist abstraction",
    "🔷 Mondrian — primary color geometric grid, De Stijl abstraction, flat bold rectangles",
    "🎸 Picasso — Cubist multiple perspectives, fragmented planes, analytical deconstruction",
    # ── SURREALISMO ─────────────────────────────────────────────
    "🌂 Magritte — surrealist hyper-realism, dreamlike impossible juxtapositions, Belgian surrealism",
    "🌀 Dalì — melting surrealist dreamscape, hyper-detailed hallucination, elongated figures",
    "🏛️ De Chirico — metaphysical painting, long dramatic shadows, empty piazzas, eerie stillness",
    # ── CONTEMPORANEO & POP ─────────────────────────────────────
    "🖌️ Banksy — urban stencil street art, black and white spray paint, sharp political irony",
    "💥 Lichtenstein — bold pop art comic book dots, primary colors, thick black outlines, Ben-Day dots",
    "🌿 Mucha — Art Nouveau decorative elegance, floral borders, soft pastel tones, ornate frames",
    "🌇 Hopper — American realism, cinematic loneliness, harsh directional light, quiet urban stillness",
    "👑 Basquiat — neo-expressionist raw energy, crown motifs, layered text and symbols, urban rawness",
    "🖍️ Scarabocchio — clumsy mouse-drawn scribble, MS Paint aesthetic, pathetically bad on purpose, white background",
]

ARTISTIC_STYLE_PROMPTS = {
    # ── RINASCIMENTO & CLASSICI ─────────────────────────────────
    "🖼️ Leonardo": (
        "Reinterpret the subject and outfit described above in the style of Leonardo da Vinci: "
        "sfumato technique with soft imperceptible tonal transitions, atmospheric depth, warm Renaissance palette. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: warm ochre, ivory, soft amber, deep brown shadows. Solid colors only, no patterns. "
        "Background: misty Renaissance landscape or dark atmospheric void. "
        "Masterwork quality with subtle painted texture."
    ),
    "🕊️ Raffaello": (
        "Reinterpret the subject and outfit described above in the style of Raphael: "
        "harmonious Renaissance composition, soft idealized grace, balanced classical symmetry. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: warm rose, sky blue, ivory, soft gold. Solid colors only, no patterns. "
        "Background: classical arch or serene landscape. "
        "Gentle diffused lighting, polished Renaissance finish."
    ),
    "⚡ Michelangelo": (
        "Reinterpret the subject and outfit described above in the style of Michelangelo: "
        "monumental sculptural presence, dramatic musculature and form, Sistine Chapel grandeur. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: warm terracotta, ivory, deep shadow tones. Solid colors only, no patterns. "
        "Background: classical architecture or dramatic void. "
        "Fresco-like painted quality, powerful chiaroscuro."
    ),
    "🕯️ Caravaggio": (
        "Reinterpret the subject and outfit described above in the style of Caravaggio: "
        "extreme chiaroscuro tenebrism, single dramatic light source cutting through absolute darkness, "
        "hyper-realistic Baroque drama. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: deep black shadows, warm candlelight amber, vivid accent colors. Solid colors only. "
        "Background: near total darkness with dramatic spotlight. "
        "Ultra-detailed painterly realism."
    ),
    # ── IMPRESSIONISMO & POST ───────────────────────────────────
    "🌸 Renoir": (
        "Reinterpret the subject and outfit described above in the style of Pierre-Auguste Renoir: "
        "soft impressionist brushwork, warm dappled light, gentle atmospheric glow, joyful luminosity. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: peach, rose, warm ivory, soft green. Solid colors only, no patterns. "
        "Background: garden, outdoor terrace or soft impressionist setting. "
        "Feathery loose brushstrokes, warm sunlit atmosphere."
    ),
    "🌻 Van Gogh": (
        "Reinterpret the subject and outfit described above in the style of Vincent van Gogh: "
        "swirling expressive impasto brushstrokes, vivid saturated palette, emotional intensity, Post-Impressionist energy. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: cobalt blue, sunflower yellow, vivid orange, deep green. Solid colors only, no patterns. "
        "Background: swirling sky or expressive landscape. "
        "Thick visible brushwork, dynamic directional strokes."
    ),
    "🐟 Matisse": (
        "Reinterpret the subject and outfit described above in the style of Henri Matisse: "
        "bold flat color fields, joyful Fauvist energy, decorative simplification, vibrant palette. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: vivid red, cerulean blue, warm yellow, bright green. Solid colors only, no patterns. "
        "Background: flat decorative color or simplified interior. "
        "Loose confident outlines, pure color expression."
    ),
    "🌙 Chagall": (
        "Reinterpret the subject and outfit described above in the style of Marc Chagall: "
        "dreamlike floating figures, folkloric symbolism, luminous jewel tones, poetic unreality. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: deep cobalt blue, violet, warm gold, crimson. Solid colors only, no patterns. "
        "Background: dreamlike floating sky or village at night. "
        "Lyrical painted quality with weightless ethereal atmosphere."
    ),
    # ── MODERNISMO & ASTRATTO ───────────────────────────────────
    "✨ Klimt": (
        "Reinterpret the subject and outfit described above in the style of Gustav Klimt: "
        "gold leaf ornamental surfaces, Byzantine decorative excess, Viennese Secession opulence. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: burnished gold, ivory, deep black, jewel tones. Solid colors only, no patterns. "
        "Background: gold mosaic or dark ornamental surface. "
        "Flat decorative quality with gilded painterly texture."
    ),
    "⭕ Mirò": (
        "Reinterpret the subject and outfit described above in the style of Joan Mirò: "
        "playful biomorphic shapes, primary colors on white, Catalan surrealist abstraction, childlike energy. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: red, blue, yellow, black on white. Solid colors only, no patterns. "
        "Background: flat white or pale ground. "
        "Bold black outlines, simplified organic forms."
    ),
    "🔷 Mondrian": (
        "Reinterpret the subject and outfit described above in the style of Piet Mondrian and De Stijl: "
        "primary color geometric grid, flat bold rectangles of red, blue, yellow, black and white. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Solid colors only, no patterns. Background: white seamless or geometric grid. "
        "Flat 2D graphic quality with bold black outlines separating color blocks."
    ),
    "🎸 Picasso": (
        "Reinterpret the subject and outfit described above in the style of Pablo Picasso: "
        "Cubist multiple simultaneous perspectives, fragmented geometric planes, analytical deconstruction of form. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: muted earth tones, grey, ochre, with sharp geometric accents. Solid colors only. "
        "Background: neutral or fragmented geometric surface. "
        "Angular broken planes, multiple viewpoints in single image."
    ),
    # ── SURREALISMO ─────────────────────────────────────────────
    "🌂 Magritte": (
        "Reinterpret the subject and outfit described above in the style of René Magritte: "
        "surrealist hyper-realism, dreamlike impossible juxtapositions, objects out of context, Belgian surrealism aesthetic. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: cool grey, ivory, soft pastel tones. Solid colors only, no patterns. "
        "Background: seamless white studio or minimal dreamlike setting. "
        "Maintain photographic quality with a painted surrealist atmosphere."
    ),
    "🌀 Dalì": (
        "Reinterpret the subject and outfit described above in the style of Salvador Dalì: "
        "melting surrealist dreamscape, hyper-detailed hallucination, elongated surreal elements, Spanish surrealism. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: earthy terracotta, ochre, warm amber. Solid colors only, no patterns. "
        "Background: desert at golden hour or vast dreamlike landscape. "
        "Ultra-detailed with a painted surrealist atmosphere."
    ),
    "🏛️ De Chirico": (
        "Reinterpret the subject and outfit described above in the style of Giorgio De Chirico: "
        "metaphysical painting aesthetics, long dramatic shadows, eerie stillness, classical architecture elements, empty piazzas. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: earthy terracotta, warm ivory, cool grey. Solid colors only, no patterns. "
        "Background: ancient ruins or empty classical piazza with dramatic shadows. "
        "Wide angle full body composition preferred."
    ),
    # ── CONTEMPORANEO & POP ─────────────────────────────────────
    "🖌️ Banksy": (
        "Reinterpret the subject and outfit described above as urban stencil art in the style of Banksy: "
        "black and white spray paint aesthetic, sharp political and artistic irony, graffiti on a brick wall or urban surface. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: monochromatic black and white with a single bold red accent element. "
        "Solid colors only. Background: urban brick wall, street setting. "
        "Stencil spray paint texture, rough edges, authentic street art feel."
    ),
    "💥 Lichtenstein": (
        "Reinterpret the subject and outfit described above in the style of Roy Lichtenstein: "
        "bold pop art comic book aesthetic, Ben-Day halftone dots, thick black outlines, primary color fills. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: flat red, yellow, blue, black and white. Solid colors only, no gradients. "
        "Background: solid color or simple geometric. "
        "Graphic comic panel quality, speech bubble optional, no text."
    ),
    "🌿 Mucha": (
        "Reinterpret the subject and outfit described above in the style of Alphonse Mucha: "
        "Art Nouveau decorative elegance, intricate floral and botanical borders, ornate circular frame, "
        "soft pastel tones with gold accents. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: soft blush, sage green, ivory, warm gold. Solid colors only, no patterns. "
        "Background: decorative Art Nouveau panel with botanical motifs. "
        "Poster-quality illustration, graceful flowing lines."
    ),
    "🌇 Hopper": (
        "Reinterpret the subject and outfit described above in the style of Edward Hopper: "
        "American realist painting, harsh directional light creating long shadows, cinematic stillness, "
        "quiet urban or architectural setting, melancholic solitude. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: warm amber light against cool blue shadow, muted earth tones. Solid colors only. "
        "Background: diner window, hotel room, empty street or sunlit building facade. "
        "Cinematic wide framing, painterly realism."
    ),
    "👑 Basquiat": (
        "Reinterpret the subject and outfit described above in the style of Jean-Michel Basquiat: "
        "neo-expressionist raw energy, crown motifs, layered gestural marks, urban rawness, "
        "primitive drawn quality over flat color backgrounds. "
        "Preserve the subject's appearance and outfit exactly as described above. "
        "Palette: raw black, vivid red, yellow, white. Solid colors only, no patterns. "
        "Background: raw canvas texture or urban wall surface. "
        "Energetic rough lines, visible reworking, expressive mark-making."
    ),
    # ── SCARABOCCHIO ────────────────────────────────────────────
    "🖍️ Scarabocchio": (
        "Redraw the subject and scene described above in the most clumsy, scribbled and pathetically bad way possible. "
        "White background. Drawn with a mouse in an old MS Paint-style program. "
        "Loosely similar to the described subject but wrong, crooked and confusing. "
        "Rough pixel-by-pixel texture, shaky unsteady lines, awkward proportions, "
        "colors slightly off, shapes vaguely recognizable but embarrassingly distorted. "
        "No artistic quality whatsoever — the charm IS the failure."
    ),
}

# Categorie per il menu artisti
ARTIST_CATEGORIES = [
    ("rinascimento", "🏛 Rinascimento & Classici", ["🖼️ Leonardo", "🕊️ Raffaello", "⚡ Michelangelo", "🕯️ Caravaggio"]),
    ("impressionismo", "🌸 Impressionismo & Post",  ["🌸 Renoir", "🌻 Van Gogh", "🐟 Matisse", "🌙 Chagall"]),
    ("modernismo", "🔷 Modernismo & Astratto",      ["✨ Klimt", "⭕ Mirò", "🔷 Mondrian", "🎸 Picasso"]),
    ("surrealismo", "🌀 Surrealismo",               ["🌂 Magritte", "🌀 Dalì", "🏛️ De Chirico"]),
    ("contemporaneo", "💥 Contemporaneo & Pop",     ["🖌️ Banksy", "💥 Lichtenstein", "🌿 Mucha", "🌇 Hopper", "👑 Basquiat"]),
    ("scarabocchio",   "🖍️ Scarabocchio",                ["🖍️ Scarabocchio"]),
]

# Mappa rapida: cat_key → (label, [artisti])
ARTIST_CAT_MAP = {key: (label, artists) for key, label, artists in ARTIST_CATEGORIES}

def get_artist_cat_keyboard():
    """Menu categorie artisti — primo livello di Stile Artistico."""
    markup = types.InlineKeyboardMarkup()
    for cat_key, cat_label, _ in ARTIST_CATEGORIES:
        markup.add(types.InlineKeyboardButton(cat_label, callback_data=f"artcat_{cat_key}"))
    markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
    return markup

def get_artist_list_keyboard(cat_key):
    """Menu artisti per una categoria specifica."""
    _, cat_label, artists = next(t for t in ARTIST_CATEGORIES if t[0] == cat_key)
    markup = types.InlineKeyboardMarkup()
    for i in range(0, len(artists), 2):
        row = [types.InlineKeyboardButton(a, callback_data=f"artist_{a}") for a in artists[i:i+2]]
        markup.row(*row)
    markup.add(types.InlineKeyboardButton("◀️ Categorie", callback_data="back_artistic"))
    return markup

def get_artist_keyboard():
    """Alias per compatibilità — mostra menu categorie artisti."""
    return get_artist_cat_keyboard()

    return markup

def build_artistic_style_prompt(artist_key=None):
    """Costruisce il prompt per il filtro artistico.
    Se artist_key è fornito usa quell'artista, altrimenti sceglie casuale."""
    if artist_key and artist_key in ARTISTIC_STYLE_PROMPTS:
        return ARTISTIC_STYLE_PROMPTS[artist_key], artist_key
    style = random.choice(ARTISTIC_STYLES_POOL)
    key = next((k for k in ARTISTIC_STYLE_PROMPTS if k in style), None)
    if key:
        return ARTISTIC_STYLE_PROMPTS[key], style
    return ARTISTIC_STYLE_PROMPTS["🌂 Magritte"], style

# ============================================================
# FILTRI
# ============================================================

# --- Y2K POP COLLAGE — POSE POOL ---
Y2K_POSE_POOL = [
    # ── FACE & HANDS ───────────────────────────────────────────
    "close-up, both hands pressed to cheeks, mouth wide open in exaggerated shock, wide eyes, direct camera gaze",
    "close-up, fingers framing face like a picture frame, bratty smirk, one eye half-closed",
    "close-up, hands covering mouth in playful secret gesture, eyes sparkling mischievous",
    "close-up, chin resting on both fists, elbows forward, bored-chic unimpressed stare",
    "close-up, index fingers pointing to cheeks with forced cute pout, exaggerated kawaii expression",
    "close-up, hands pulling hair dramatically, mouth open in playful distress, theatrical expression",
    # ── CROUCHING & LOW ─────────────────────────────────────────
    "full body crouching low, knees folded inward pigeon-toed, arms loosely draped over knees, eyes dramatically closed",
    "deep squat, arms wrapped around knees, chin tucked, looking up at camera with wide eyes",
    "crouching with one knee on floor, other leg extended sideways, hand resting on raised knee, seductive direct gaze",
    "low crouch, both hands flat on floor between feet, head lifted defiantly toward camera",
    "crouching with elbows on knees, hands dangling, head tilted sideways with playful smirk",
    # ── STANDING DYNAMIC ────────────────────────────────────────
    "full body standing, one hand on hip, other hand flicking hair, slight weight shift, confident sporty pose",
    "standing with legs wide apart, both arms crossed over chest, powerful stance, intense stare",
    "standing twisting torso sharply to one side, one arm pointing toward camera, other arm back",
    "standing on one leg, other knee raised to hip height, arms out for balance, playful energy",
    "full body standing, back to camera, head turned all the way over shoulder with direct eye contact",
    "standing, both arms raised overhead making peace signs with both hands, big bright smile",
    # ── LENS DISTORTION & DYNAMIC PERSPECTIVE ───────────────────
    "dynamic wide-angle distortion, one arm and hand fully extended toward camera lens creating exaggerated perspective, body receding dramatically behind",
    "leaning sharply forward toward lens, face close to camera, hands braced on thighs, confrontational editorial energy",
    "low camera angle shooting upward, subject standing tall, one hand on hip, powerful towering perspective",
]

FILTERS = {

    # ── STILISTICI ──────────────────────────────────────────
    "cinematic_highangle": {
        "label": "⬆️⬇️ Cinematic High-Angle",
        "cat": "stylistic",
        "prompt": (
            "Generate a cinematic high-angle portrait of the subject and outfit described above. "
            "Camera positioned overhead looking down, intense expressive eyes with sharp focus. "
            "Shallow depth of field, softly blurred background, dramatic soft lighting, moody color grading, high contrast. "
            "Professional fashion photography, 85mm lens, f/1.8, cinematic realism, editorial portrait, 8K, film grain."
        )
    },
    "dramatic": {
        "label": "⬆️ Dramatic Low-Angle",
        "cat": "stylistic",
        "prompt": (
            "Generate a dramatic low-to-high angle shot of the subject and outfit described above. "
            "Camera positioned low, shooting upward. Extreme contrast, rich deep colors, especially reds and blacks. "
            "Darkened background, bright direct theatrical lighting. "
            "High-fashion editorial, cinematic, 8K."
        )
    },
    "glossy": {
        "label": "🌟 Glossy Opal",
        "cat": "stylistic",
        "prompt": (
            "Generate a hyper-glamour image of the subject and outfit described above. "
            "Apply Ultra-Opal Fairy-Angel Couture style: 3D iridescent aesthetics, multi-layered pearlescent reflections, "
            "rainbow opalescence, liquid glows, Swarovski crystals, mirror-gloss surfaces. "
            "Lighting: soft golden-pink atmospheric scattering, suspended micro-sparkles, precious bokeh. "
            "Iridescent shifting reflections only on the outfit — pearl-champagne tone, cold-pink hot spots, blue-opal touches. "
            "Skin untouched. 8K editorial."
        )
    },
    "iridescent": {
        "label": "🌈 Iridescent",
        "cat": "stylistic",
        "prompt": (
            "Generate an iridescent editorial image of the subject and outfit described above. "
            "Dichroic refraction creating a spectrum of shifting colors: electric teal, deep cosmic blue, fiery metallic orange. "
            "All non-skin surfaces have pearlescent glow, high-gloss textures, micro-crystalline details. "
            "Prismatic light dispersion, volumetric studio lighting, holographic effect on outfit and accessories. "
            "Skin realistic and untouched. Color palette: burning gold and vibrant cyan. Ultra-high definition, 8K, cinematic fashion."
        )
    },
    "galaxy": {
        "label": "🌌 Galaxy Couture",
        "cat": "stylistic",
        "prompt": (
            "Generate a jewel galaxy haute couture editorial of the subject and outfit described above. "
            "Palette: midnight blue, purple, turquoise, warm gold, soft iridescent nebula tones. "
            "Brilliant metallic reflections and colored edge glints on the outfit as if coated in iridescent metal. "
            "Small gem and crystal-like lights in amber-orange, aquamarine, opalescent white as decorative details. "
            "Enhanced contrast and sharpness. Do not modify shapes — only colors and textures. Fashion cover quality, 8K."
        )
    },
    "arabesque": {
        "label": "🌹 Arabesque",
        "cat": "stylistic",
        "prompt": (
            "Generate a baroque maximalist editorial of the subject and outfit described above. "
            "Every surface covered in ornament: intricate gold filigree, Sicilian baroque motifs, Byzantine gold mosaics, "
            "maiolica tile patterns in cobalt and gold, Versailles excess, Venetian carnival richness. "
            "The outfit transformed into the most encrusted jewel-laden embroidered beaded and gold-threaded garment imaginable. "
            "Skin and face remain photorealistic and untouched. "
            "Palette: deep crimson, imperial gold, byzantine cobalt, emerald green, ivory, black. "
            "Ultra-detailed 8K, dramatic chiaroscuro lighting."
        )
    },

    # ── FANTASY & ART ────────────────────────────────────────
    "stained_glass": {
        "label": "🎐 Stained Glass",
        "cat": "fantasy",
        "prompt": (
            "Transform the subject and outfit described above into a hyper-realistic figure "
            "made entirely of stained glass and translucent crystal. "
            "The whole figure is a mosaic of milky white and pale blue glass tiles held together by polished silver wire frame. "
            "Hair flows in sculpted waves of transparent teal glass. "
            "Soft cinematic lighting creating brilliant specular highlights and internal refractions. "
            "Dark moody background, 8K resolution, 85mm lens, macro photography detail, iridescent textures."
        )
    },
    "underwater": {
        "label": "🧜 Underwater Gold",
        "cat": "fantasy",
        "prompt": (
            "Generate a high-fantasy underwater image of the subject and outfit described above. "
            "Style: high-brilliance gold and saturated turquoise and teal palette, volumetric light from above, "
            "translucent flowing fabrics like silk, hyper-detailed jewelry in gold and turquoise stones, "
            "thousands of micro-bubbles and suspended particles, mystical atmosphere in deep blue-cyan water, "
            "ultra-detailed glossy high-contrast rendering with specular reflections."
        )
    },
    "underwater_editorial": {
        "label": "🤿 Underwater Editorial",
        "cat": "fantasy",
        "prompt": (
            "Generate a cinematic underwater fashion editorial. The subject described above is submerged "
            "in crystal-clear tropical water — turquoise and gold-lit from above by filtered sunlight. "
            "Surrounded by golden light rays, tropical fish, coral or floating fabric. "
            "The subject floats or poses gracefully underwater, hair and fabric moving in slow underwater motion. "
            "Faithfully preserve the exact outfit from the reference, adapted to underwater movement. "
            "Lighting: underwater caustic light — shimmering golden sun rays filtering through the water surface, "
            "warm golden and turquoise tones, luminous and ethereal atmosphere. "
            "Cinematic wide-to-medium shot, subject sharp against soft blue-gold water bokeh. "
            "Dreamy, ethereal, high-fashion. 8K ultra-realistic. "
            "NEGATIVE PROMPT — EXTRA: body hair, chest hair, arm hair, leg hair, dry setting, studio background, indoor."
        )
    },
    "3d_synthetic": {
        "label": "🪟 3D Synthetic",
        "cat": "fantasy",
        "prompt": (
            "Generate a hyper-realistic cutting-edge 3D rendering of the subject and outfit described above. "
            "Transform all surfaces into high-quality translucent synthetic with dazzling glossy multichromatic finish "
            "and subtle knurled texture refracting light into razor-sharp specular reflections. "
            "Multi-source HDRI studio lighting with intense rim and backlighting. "
            "Colors: electric cyan, deep violet, molten gold gradient. "
            "Pronounced bloom effects, chromatic aberration. Absolute black background, "
            "flat frontal eye-level macro perspective, aggressive contrast, saturated color grading, subtle digital grain."
        )
    },
    "graffiti": {
        "label": "🧯 Graffiti Artist",
        "cat": "fantasy",
        "prompt": (
            "Generate an ultra-realistic image of the subject described above "
            "spray-painting a full-body self-portrait graffiti on an urban brick wall. "
            "The subject wears the exact outfit described above. "
            "The subject is standing or half-crouched mid-spray while the full mural is clearly visible on a continuous brick wall. "
            "Fresh graffiti with natural overspray and paint drips, spray mist visible, "
            "several unbranded spray cans on the ground. Light paint speckles on clothes or shoes only. "
            "Realistic urban daylight, shallow depth of field, high-resolution, no extra people, no text."
        )
    },
    "cloud_sculpture": {
        "label": "☁️ Cloud Sculpture",
        "cat": "fantasy",
        "use_master": False,
        "prompt": (
            "Transform the subject described above into a form made entirely of soft billowing clouds. "
            "Preserve the overall silhouette and posture. "
            "Every part — body, clothing, accessories, hair — fully transformed into cloud. "
            "No original colors or textures remain. Pure cloud only: fluffy rounded voluminous masses of white vapor. "
            "No sharp edges, no fine surface detail — only soft rounded cloud puffs merging into each other. "
            "Facial features suggested by cloud volume shapes and shadows only. Form dissolves softly at edges into sky. "
            "Lighting: warm natural sunlight from above with gentle rounded shadows. "
            "Background: bright open blue sky. Airy weightless dreamlike mood. Ultra-realistic cloud render quality."
        )
    },
    "artistic_style": {
        "label": "🎨 Stile Artistico",
        "cat": "artistic",
        "use_master": False,
        "prompt": "__artistic_style__",
    },
    "lego": {
        "label": "🧱 LEGO",
        "cat": "lego",
        "prompt": (
            "Generate an ultra-photorealistic LEGO brick-built version of the subject and outfit "
            "described above. "
            "Constructed entirely from authentic LEGO bricks, plates, tiles, slopes, studs and technic elements. "
            "Highly detailed studded surfaces, visible seams between bricks, layered LEGO geometry, "
            "glossy ABS plastic material with subtle scratches and fingerprints, accurate plastic reflections. "
            "Vibrant molded colors, micro surface imperfections. "
            "Dynamic cinematic composition with LEGO bricks exploding and assembling around the subject. "
            "Professional studio lighting, softbox key light, rim light, dramatic contrast, global illumination, ray-traced reflections."
        )
    },
    "dissolve": {
        "label": "💧 Dissolvence",
        "cat": "stylistic",
        "use_master": False,
        "prompt": (
            "Generate a portrait of the subject described above dissolving into fluid motion trails. "
            "The subject wears the exact outfit described above. "
            "Soft painterly smears extending from contours, flowing movement captured in long exposure, "
            "elegant minimal background. "
            "Camera: SONY VENICE, ARRI Signature Prime 50mm. "
            "Lighting: soft studio lighting, luminous highlights, gentle directional light. "
            "Motion: long exposure smear, flowing liquid-like blur trails, continuous movement distortion. "
            "Color: soft neutral tones with subtle color bleed. Surreal editorial portrait, cinematic fine art photography."
        )
    },
    "ghost_temporal": {
        "label": "👻 Ghost Temporal",
        "cat": "stylistic",
        "use_master": False,
        "prompt": (
            "Generate an ultra-realistic 8K cinematic studio portrait of the subject and outfit "
            "described above, framed from mid-thigh up. "
            "Background: vibrant ochre-red (#C0392B), uniform but subtly graded for depth. "
            "Pose: subject standing confidently, sharp focus, primary figure in full presence. "
            "Ghost effect: a translucent motion-blurred ghost duplicate of the subject positioned slightly behind and to the right, "
            "streaking horizontally with colorful light trails (red, blue, yellow) conveying rapid movement or temporal distortion. "
            "The ghost is semi-transparent 40-50% opacity with chromatic aberration streaks. "
            "Lighting: harsh frontal studio lighting, crisp shadows, emphasizing fabric textures. "
            "High-fashion editorial style, shallow depth of field on primary figure, 85mm lens feel."
        )
    },
    "long_exposure": {
        "label": "📸 Long Exposure",
        "cat": "stylistic",
        "use_master": False,
        "prompt": (
            "Recreate the subject and outfit described above as a long exposure fashion photograph. "
            "Technique: slow shutter speed 1-4 seconds, camera on tripod, subject in deliberate motion. "
            "The face and upper torso retain sharpness while body, arms and hair streak into directional motion blur trails. "
            "Multiple semi-transparent ghost overlaps of the same figure suggest continuous movement. "
            "Background environment remains completely static and sharp — high contrast with blurred subject. "
            "Lighting: ambient light sources sharp, their light bleeds into motion trails creating warm luminous streaks. "
            "Mood: ethereal, ghostly, cinematic fine art. "
            "Technical: 8K, 35-50mm lens, f/8, ISO 400, natural motion smear physics — no artificial radial blur."
        )
    },

    # ── SCENOGRAFICI ─────────────────────────────────────────
    "giantess": {
        "label": "🏙️ Giantess NYC",
        "cat": "scenic",
        "prompt": (
            "Generate a hyper-realistic photograph of the subject described above "
            "as a giant hundreds of meters tall, walking carefully down the center of Broadway in New York City. "
            "The subject wears the exact outfit described above. "
            "The sneakers span several city blocks while the head towers above the spire of the Chrysler Building. "
            "Real-world location with authentic textures of glass, concrete and foliage. "
            "No toys, no diorama, no miniature effect. Sharp focus throughout. "
            "Natural daylight, realistic atmospheric perspective, 8K resolution, tilt-shift effect."
        )
    },
    "action_figure": {
        "label": "🪆 Action Figure",
        "cat": "scenic",
        "prompt": (
            "Create a 1/7 scale commercial figure of the subject described above, "
            "wearing the exact outfit described above, "
            "set on a computer desk with a transparent round acrylic base. "
            "Next to the desk appears the same subject life-sized wearing the same outfit, "
            "carefully cleaning the figure with a fine brush in a modern well-lit studio. "
            "Background: a collection of toys and action figures."
        )
    },
    "art_doll": {
        "label": "👯 Art Doll Exhibition",
        "cat": "scenic",
        "prompt": (
            "In a bright minimalist art exhibition space, generate an oversized sculpture in the style of a cute big-eyed doll, "
            "wearing clothing style, hairstyle and accessories identical to those described above. "
            "The sculpture is 50% taller than a real person and stands naturally behind and slightly to the side of the subject. "
            "The subject described above stands in the foreground — face and pose natural. "
            "Lighting: soft but well-defined. Image clear and rich in detail, creating a cute and trendy exhibition atmosphere."
        )
    },
    "toy_window": {
        "label": "🎎 Toy Store Window",
        "cat": "scenic",
        "prompt": (
            "Generate an ultra-realistic image of the subject described above "
            "standing in front of a luxury toy store window, delicately touching the glass with one hand. "
            "The subject wears the exact outfit described above. "
            "Inside the display window there is a full-size cartoon doll modeled on the subject: "
            "same features, same hair, same outfit as described above, "
            "but rendered as a cute animated character with large eyes and stylized proportions. "
            "The doll is hyper-defined with premium toy rendering. "
            "Realistic reflections on the store window, high-level fashion look. "
            "Settings: f/2.8, ISO 200, 85mm lens, soft cinematic depth of field, warm soft light, ultra-detailed glossy finish."
        )
    },
    "selfie_stick": {
        "label": "🤳 Selfie Stick POV",
        "cat": "scenic",
        "use_master": False,
        "prompt": (
            "Generate an editorial photograph in selfie stick POV style of the subject and outfit "
            "described above. "
            "Camera setup: extreme wide-angle fisheye perspective shot from above, mounted at the top of a selfie stick "
            "held by the subject RIGHT hand. The stick extends upward from the hand and exits the frame at the top. "
            "The LEFT arm hangs naturally at the side. Strong fisheye barrel distortion on the edges. "
            "The subject looks up toward the lens with a natural smile. "
            "Framing: bird's eye perspective — subject seen from 2-3 meters above, environment extends dramatically in all directions. "
            "Scene: one of these environments at random: wide sandy beach with waves, blooming wildflower meadow, "
            "colorful city street, mountain trail. "
            "Technical: GoPro aesthetic, vivid saturated colors, high dynamic range, sharp foreground, wide depth of field. "
            "Natural outdoor lighting. Cinematic 1:1 aspect ratio. Ultra-realistic, no text, no watermark."
        )
    },

    # ── COLLAGE ───────────────────────────────────────────────
    "new_pose": {
        "label": "🆕 New Pose",
        "cat": "collage",
        "prompt": (
            "Generate a new image of the subject and outfit described above "
            "with the exact same scene and setting but with a completely different natural editorial pose "
            "and a different facial expression."
        )
    },
    "triple_set": {
        "label": "3️⃣× Triple Set",
        "cat": "collage",
        "prompt": (
            "Generate a set of 3 separate high-resolution images of the subject and outfit "
            "described above. "
            "Each image has a completely different pose. "
            "Maintain original aspect ratio, exact proportions, same setting and lighting across all 3."
        )
    },
    "triptych": {
        "label": "3️⃣❌1️⃣ Triptych GHI",
        "cat": "altri",
        "prompt": (
            "Generate a 3:1 collage of the subject and outfit described above, "
            "with three completely different poses and facial expressions. "
            "Version 1 — Glossy: hyper-glamour rendering, pearlescent iridescent aesthetics, mirror-gloss surfaces, angel couture runway effect. "
            "Version 2 — HDR+: powerful neon light sources along hair, clothes and background edges, "
            "cyan, magenta and lime glow, HDR long-exposure neon style, face and eyes sharp and realistic. "
            "Version 3 — Holographic: dichroic refraction spectrum shifting between electric teal, deep cosmic blue "
            "and fiery metallic orange, prismatic light dispersion, holographic shimmer on all non-skin surfaces."
        )
    },
    "emoji_art": {
        "label": "🧱 LEGO Mosaic",
        "cat": "lego",
        "use_master": False,
        "prompt": "__emoji_art__",
    },
    "lego_galaxy": {
        "label": "🌌 LEGO Galaxy",
        "cat": "lego",
        "use_master": False,
        "prompt": "__lego_galaxy__",
    },
    "pastel_clones": {
        "label": "🌸 Pastel Clones",
        "cat": "collage",
        "use_master": False,
        "prompt": (
            "Generate an ultra-realistic high-fashion editorial collage with 7 clones of the subject "
            "described above. "
            "Each clone wears the exact outfit described above — "
            "same garments, same colors, same cut, same accessories. "
            "Pastel treatment applies ONLY to: background colors, studio lighting palette, props, set elements, "
            "floor surfaces, decorative objects. "
            "Palette: blush pink, powder blue, mint green, lilac, cream — soft geometric color block background panels. "
            "The subject outfit and skin tones are NOT pastelized — they remain true to the subject description above. "
            "Layout: luxury fashion magazine spread, structured layered depth. "
            "7 poses: standing, seated, lying, walking, cross-legged, playful, relaxed. "
            "Layered torn paper collage textures, floating petals. "
            "Soft diffused beauty lighting with subtle rim light, cinematic glow. "
            "85mm portrait lens, ultra-detailed, ratio 9:16."
        )
    },
    "collage_2x2": {
        "label": "🟦 Collage 2×2",
        "cat": "collage",
        "use_master": False,
        "prompt": (
            "Generate a 2x2 quad-panel collage of the subject and outfit described above, "
            "with four completely different camera angles and poses. "
            "All four panels must show the same subject with the exact outfit from the subject description above consistently. "
            "Top-left: camera positioned low angled upward, subject has one hand on hip the other extended toward the lens. "
            "Top-right: extreme high-angle bird's eye view looking straight down, subject gazes up at camera one arm reaching upward. "
            "Bottom-left: full-length shot with tilted horizon, subject in three-quarter profile highlighting silhouette. "
            "Bottom-right: shot from behind and slightly to the side, subject looks back over the shoulder toward the camera. "
            "Thin dark dividing lines between panels. Unified cinematic look."
        )
    },
    "photobooth_4x4": {
        "label": "📷 Photobooth 4×4",
        "cat": "collage",
        "use_master": False,
        "prompt": (
            "Generate a color photobooth expression grid of the subject and outfit described above, "
            "with a 4x4 layout of 16 panels. "
            "Replicate all styling details consistently across every single panel. "
            "Setting: tight head-and-shoulders photobooth framing, 50mm lens look. "
            "Eyes sharp in every panel. Natural realistic skin tones. Medium to high contrast. Thin gutters between panels. "
            "16 expressions in order: 1-scrunched smile eyes slightly squeezed, 2-intense stare fingers framing eyes, "
            "3-big joyful laugh mouth open, 4-bored unimpressed chin in hands, 5-sad pout watery eyes, "
            "6-goofy face hands making small horns above head, 7-playful tongue out cheeky grin, "
            "8-angry glare eyebrows down, 9-flirty look hand touching cheek, 10-surprised wide eyes mouth slightly open, "
            "11-excited shout hands near face, 12-mischievous grin claw-like hand pose, "
            "13-confused frown lips pressed, 14-dramatic crying hands on head, "
            "15-tongue out eyes closed playful, 16-duck face with small devil horns gesture. "
            "Ultra high resolution photorealistic color photobooth contact sheet. "
            "No extra fingers, no deformed hands, no text, no watermark, no blur."
        )
    },
    "fullbody_3x3": {
        "label": "🧍 Full Body 3×3",
        "cat": "collage",
        "use_master": False,
        "prompt": (
            "Generate a full body editorial grid of the subject and outfit described above, "
            "with exactly 3 columns and 3 rows — 9 panels total. "
            "Each panel shows the full body from head to toe — outfit fully visible in every panel. "
            "Natural realistic skin tones. Medium to high contrast. Thin gutters between panels. "
            "High panel consistency. 85mm lens look. Cinematic editorial quality. "
            "9 full body poses: "
            "1-low angle shot from below dramatic perspective subject looking down at camera, "
            "2-arms raised overhead full body stretch triumphant pose, "
            "3-crouching low dramatic editorial crouch arms resting on knees, "
            "4-crossed arms power stance slight smirk direct gaze, "
            "5-side profile elegant posture chin slightly raised, "
            "6-mid-turn caught in movement hair and outfit in motion, "
            "7-leaning forward toward camera hands on thighs intimate editorial, "
            "8-seated on floor legs extended arms behind relaxed editorial pose, "
            "9-full rear view looking back at camera. "
            "Ultra high resolution photorealistic editorial fashion contact sheet."
        )
    },

    "y2k_pop_collage": {
        "label": "🌟 Y2K Pop Collage",
        "cat": "collage",
        "use_master": False,
        "prompt": "__y2k_pop_collage__",
    },

    # ── PET ──────────────────────────────────────────────────────
    "pet_mosaic": {
        "label": "🐾 Pet Mosaic 4×4",
        "cat": "altri",
        "use_master": False,
        "prompt": (
            "Generate a color expression grid of the animal described above, "
            "with exactly 4 columns and 4 rows — 16 panels total. "
            "Preserve the exact animal: same breed, fur color, fur texture, eye color, size, markings. "
            "Same animal in every single panel. "
            "Setting: plain neutral light gray backdrop, soft even studio lighting, tight head and shoulders framing. "
            "50mm lens look. Eyes sharp in every panel. Natural realistic fur texture. Medium to high contrast. "
            "Thin gutters between panels. "
            "16 expressions: 1-happy panting tongue out wide smile, 2-ears perked up alert and curious staring at camera, "
            "3-head tilted sideways maximum confusion, 4-yawning big mouth open eyes squinting, "
            "5-sleepy half-closed eyes drowsy expression, 6-playful bow front legs down ready to play, "
            "7-grumpy serious unimpressed face, 8-surprised wide eyes startled expression, "
            "9-sniffing nose forward intense concentration, 10-excited bouncy energy sparkling eyes, "
            "11-begging puppy eyes maximum cuteness, 12-caught doing something naughty guilty face ears back, "
            "13-dramatic sad face watery eyes pout, 14-fierce mock growl showing teeth playfully, "
            "15-blep tiny tongue barely sticking out, 16-ridiculous goofy face with small devil horns gesture. "
            "Ultra high resolution photorealistic animal portrait contact sheet. No text, no watermark, no blur."
        )
    },
    "lego_minifig": {
        "label": "🟡 LEGO Minifig",
        "cat": "lego",
        "use_master": False,
        "prompt": "__lego_minifig__",
    },
}

CATEGORIES = {
    "stylistic": "🎨 Stilistici",
    "fantasy":   "✨ Fantasy & Art",
    "scenic":    "🏙️ Scenografici",
    "collage":   "🖼️ Collage",
    "mosaic":    "🖼️ Mosaic",
    "artistic":  "🎨 Stile Artistico",
    "lego":      "🧱 LEGO",
    "altri":     "✨ Altri",
}

def filters_by_cat(cat):
    return {k: v for k, v in FILTERS.items() if v["cat"] == cat}

# ============================================================
# BOT
# ============================================================
telebot.apihelper.ENABLE_MIDDLEWARE = True
bot = telebot.TeleBot(TOKEN, parse_mode="HTML", use_class_middlewares=True)

# --- WHITELIST MIDDLEWARE ---
@bot.middleware_handler(update_types=['message', 'callback_query'])
def check_whitelist(bot_instance, update):
    if hasattr(update, 'from_user'):
        uid = update.from_user.id
    elif hasattr(update, 'message') and update.message:
        uid = update.message.from_user.id
    else:
        return
    if not is_allowed(uid):
        logger.warning(f"🚫 Accesso non autorizzato bloccato: uid={uid}")
        raise telebot.CancelUpdate()

# --- KEYBOARD CATEGORIE ---
def cat_keyboard():
    markup = types.InlineKeyboardMarkup()
    for cat_key, cat_label in CATEGORIES.items():
        if cat_key == "mosaic":
            markup.add(types.InlineKeyboardButton(cat_label, callback_data="mosaic_start"))
        elif cat_key == "artistic":
            markup.add(types.InlineKeyboardButton(cat_label, callback_data="cat_artistic"))
        elif cat_key == "lego":
            markup.add(types.InlineKeyboardButton(cat_label, callback_data="cat_lego"))
        else:
            markup.add(types.InlineKeyboardButton(cat_label, callback_data=f"cat_{cat_key}"))
    return markup

def artist_cat_keyboard():
    """Sottomenu categorie artisti — primo livello."""
    markup = types.InlineKeyboardMarkup()
    for cat_label, _ in ARTIST_CATEGORIES:
        markup.add(types.InlineKeyboardButton(cat_label, callback_data=f"artcat_{cat_label}"))
    markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
    return markup

def artist_list_keyboard(cat_label):
    """Lista artisti di una categoria — secondo livello."""
    markup = types.InlineKeyboardMarkup()
    artists = next((a for l, a in ARTIST_CATEGORIES if l == cat_label), [])
    for i in range(0, len(artists), 2):
        row = [types.InlineKeyboardButton(a, callback_data=f"artist_{a}") for a in artists[i:i+2]]
        markup.row(*row)
    markup.add(types.InlineKeyboardButton("◀️ Categorie", callback_data="cat_artistic"))
    return markup

# --- KEYBOARD FILTRI ---
def filter_keyboard(cat):
    markup = types.InlineKeyboardMarkup()
    for fkey, fdata in filters_by_cat(cat).items():
        markup.add(types.InlineKeyboardButton(fdata["label"], callback_data=f"f_{fkey}"))
    markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
    return markup

# --- /start ---
@bot.message_handler(commands=['start', 'reset'])
def cmd_start(m):
    uid = m.from_user.id
    if not is_allowed(uid):
        logger.warning(f"🚫 /start non autorizzato: uid={uid} username={m.from_user.username}")
        return
    username = m.from_user.username or m.from_user.first_name
    # Reset completo di tutti gli stati
    user_filter[uid] = None
    user_category[uid] = None
    user_artist.pop(uid, None)
    pending.pop(uid, None)
    last_img.pop(uid, None)
    last_prompt.pop(uid, None)
    pending_artistic_style.pop(uid, None)
    stereo_waiting.pop(uid, None)
    stereo_last_img.pop(uid, None)
    mirror_waiting.pop(uid, None)
    mirror_last_img.pop(uid, None)
    mirror_pending_prompt.pop(uid, None)
    if uid in mosaic_collecting:
        session = mosaic_collecting.pop(uid)
        if session.get('timer'):
            try: session['timer'].cancel()
            except Exception: pass
    logger.info(f"🔄 /start da {username} (id={uid})")
    bot.send_message(m.chat.id, "✅ <b>Reset completo.</b> Tutte le sessioni cancellate.")
    bot.send_message(m.chat.id,
        f"<b>🎨 Filtro v{VERSION}</b>\n\n"
        f"Invia una foto e scegli un filtro.\n"
        f"Usa /filtro per scegliere il filtro prima della foto, o invia la foto direttamente.",
        reply_markup=cat_keyboard()
    )

# --- /filtro ---
@bot.message_handler(commands=['filtro', 'filter'])
def cmd_filtro(m):
    uid = m.from_user.id
    logger.info(f"🎨 /filtro da {m.from_user.username or m.from_user.first_name} (id={uid})")
    bot.send_message(m.chat.id, "🎨 <b>Scegli una categoria:</b>", reply_markup=cat_keyboard())

# --- /help ---
@bot.message_handler(commands=['help'])
def cmd_help(m):
    bot.send_message(m.chat.id,
        f"<b>🎨 Filtro v{VERSION} — Comandi</b>\n\n"
        f"/start · /reset — Reset e menu principale\n"
        f"/filtro — Scegli filtro prima della foto\n"
        f"/mosaic — Avvia sessione mosaic (2-9 foto)\n"
        f"/done — Conclude sessione mosaic\n"
        f"/caption — Genera caption da foto\n"
        f"/lastprompt — Ultimo prompt inviato\n"
        f"/info — Versione e informazioni\n"
        f"/help — Questo messaggio\n\n"
        f"<i>Invia una foto per iniziare.</i>"
    )

# --- /info ---
@bot.message_handler(commands=['info'])
def cmd_info(m):
    uid = m.from_user.id
    fkey = user_filter[uid]
    fname = FILTERS[fkey]["label"] if fkey else "Nessuno"
    bot.send_message(m.chat.id,
        f"<b>🎨 Filtro v{VERSION}</b>\n\n"
        f"Filtro attivo: <b>{fname}</b>\n"
        f"Categorie: Stilistici · Fantasy & Art · Scenografici · Collage · Mosaic · Altri\n"
        f"Comandi: /help per la lista completa"
    )

@bot.message_handler(commands=['lastprompt'])
def cmd_lastprompt(m):
    uid = m.from_user.id
    username = m.from_user.username or m.from_user.first_name
    logger.info(f"🔍 /lastprompt da {username} (id={uid})")
    prompt = last_prompt.get(uid)
    if not prompt:
        bot.send_message(m.chat.id, "⚠️ Nessun prompt disponibile. Genera prima un'immagine.")
        return
    chunks = [prompt[i:i+3800] for i in range(0, len(prompt), 3800)]
    for idx, chunk in enumerate(chunks):
        header = f"🔍 <b>Ultimo prompt inviato all'API</b> ({idx+1}/{len(chunks)}):\n\n" if idx == 0 else f"<i>(continua {idx+1}/{len(chunks)})</i>\n\n"
        bot.send_message(m.chat.id, f"{header}<code>{html.escape(chunk)}</code>")

# --- CALLBACK POST-GENERAZIONE ---
@bot.callback_query_handler(func=lambda c: c.data in ["post_newfilter", "post_newphoto", "post_newboth", "post_retry"])
def handle_post(call):
    uid = call.from_user.id
    username = call.from_user.username or call.from_user.first_name

    try:
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass

    if call.data == "post_retry":
        # Stesso filtro, stessa foto
        fkey = user_filter.get(uid)
        img = last_img.get(uid)
        if not fkey or not img:
            bot.send_message(call.message.chat.id, "⚠️ Nessun dato da ripetere. Invia una nuova foto.")
            return
        fname = FILTERS[fkey]["label"]
        logger.info(f"🔁 {username} → riprova stesso filtro ({fkey}), stessa foto")
        bot.send_message(call.message.chat.id, f"🔁 <b>Riprovo {fname}...</b>")
        data = {'filter_key': fkey, 'img': img}
        if fkey == "emoji_art":
            executor.submit(_run_emoji_art, call.message.chat.id, uid, username, img)
        elif fkey == "lego_galaxy":
            executor.submit(_run_lego_galaxy, call.message.chat.id, uid, username, img)
        elif fkey == "lego_minifig":
            executor.submit(_run_lego_minifig, call.message.chat.id, uid, username, img)
        else:
            executor.submit(_run_generation, call.message.chat.id, uid, username, data)
        return

    if call.data == "post_newfilter":
        # Stessa foto, nuovo filtro — resetta solo il filtro
        if uid in last_img:
            pending[uid] = {'img': last_img[uid], 'filter_key': None}
        user_filter[uid] = None
        logger.info(f"🔄 {username} → nuovo filtro, stessa foto")
        bot.send_message(call.message.chat.id, "🎨 <b>Scegli il nuovo filtro:</b>", reply_markup=cat_keyboard())

    elif call.data == "post_newphoto":
        # Stesso filtro, nuova foto
        fkey = user_filter[uid]
        if not fkey:
            bot.send_message(call.message.chat.id, "⚠️ Nessun filtro attivo. Scegli un filtro prima.", reply_markup=cat_keyboard())
            return
        fname = FILTERS[fkey]["label"]
        logger.info(f"🔄 {username} → stesso filtro ({fkey}), nuova foto")
        bot.send_message(call.message.chat.id,
            f"📷 Filtro attivo: <b>{fname}</b>\n\nInvia la nuova foto da elaborare.")

    elif call.data == "post_newboth":
        # Nuova foto E nuovo filtro — reset completo
        user_filter[uid] = None
        user_category[uid] = None
        pending.pop(uid, None)
        logger.info(f"🔄 {username} → nuova foto e nuovo filtro")
        bot.send_message(call.message.chat.id,
            "🆕 <b>Ricominciamo!</b>\n\nInvia una nuova foto e scegli un nuovo filtro.",
            reply_markup=cat_keyboard())

# --- CALLBACK CATEGORIE ---
@bot.callback_query_handler(func=lambda c: c.data.startswith("cat_") or c.data == "back_cats")
def handle_cat(call):
    uid = call.from_user.id
    if call.data == "back_cats":
        try:
            bot.edit_message_text("🎨 <b>Scegli una categoria:</b>",
                call.message.chat.id, call.message.message_id,
                reply_markup=cat_keyboard(), parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id, "🎨 <b>Scegli una categoria:</b>",
                reply_markup=cat_keyboard())
        return

    cat = call.data.replace("cat_", "")
    user_category[uid] = cat

    # Categoria "artistic" → menu categorie artisti
    if cat == "artistic":
        try:
            bot.edit_message_text("🎨 <b>Stile Artistico</b> — scegli la categoria:",
                call.message.chat.id, call.message.message_id,
                reply_markup=get_artist_cat_keyboard(), parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id, "🎨 <b>Stile Artistico</b> — scegli la categoria:",
                reply_markup=get_artist_cat_keyboard(), parse_mode="HTML")
        return

    # Categoria "lego" → submenu LEGO
    if cat == "lego":
        markup = types.InlineKeyboardMarkup()
        for fkey, fdata in filters_by_cat("lego").items():
            markup.add(types.InlineKeyboardButton(fdata["label"], callback_data=f"f_{fkey}"))
        markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
        try:
            bot.edit_message_text("🧱 <b>LEGO</b> — scegli il filtro:",
                call.message.chat.id, call.message.message_id,
                reply_markup=markup, parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id, "🧱 <b>LEGO</b> — scegli il filtro:",
                reply_markup=markup, parse_mode="HTML")
        return

    # Categoria "altri" ha anche 3D e Mirror che non sono in FILTERS
    if cat == "altri":
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🎥 3D Stereo", callback_data="stereo_start"))
        markup.add(types.InlineKeyboardButton("🪞 Mirror Outfits", callback_data="mirror_start"))
        for fkey, fdata in filters_by_cat("altri").items():
            markup.add(types.InlineKeyboardButton(fdata["label"], callback_data=f"f_{fkey}"))
        markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
        try:
            bot.edit_message_text("✨ <b>Altri</b> — scegli il filtro:",
                call.message.chat.id, call.message.message_id,
                reply_markup=markup, parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id, "✨ <b>Altri</b> — scegli il filtro:",
                reply_markup=markup)
        return

    cat_label = CATEGORIES.get(cat, cat)
    try:
        bot.edit_message_text(f"🎨 <b>{cat_label}</b> — scegli il filtro:",
            call.message.chat.id, call.message.message_id,
            reply_markup=filter_keyboard(cat), parse_mode="HTML")
    except Exception:
        bot.send_message(call.message.chat.id, f"🎨 <b>{cat_label}</b> — scegli il filtro:",
            reply_markup=filter_keyboard(cat))

# --- CALLBACK FILTRI ---
@bot.callback_query_handler(func=lambda c: c.data.startswith("f_"))
def handle_filter(call):
    uid = call.from_user.id
    fkey = call.data.replace("f_", "")
    if fkey not in FILTERS:
        bot.answer_callback_query(call.id, "Filtro non trovato.")
        return

    user_filter[uid] = fkey
    fname = FILTERS[fkey]["label"]
    bot.answer_callback_query(call.id, f"✅ {fname}")
    logger.info(f"🎨 {call.from_user.username or call.from_user.first_name} (id={uid}) → filtro: {fkey}")

    # Pulisce sessione mosaic zombie se presente
    if uid in mosaic_collecting:
        session = mosaic_collecting.pop(uid)
        if session.get("timer"):
            session["timer"].cancel()
        logger.info(f"   🧹 Sessione mosaic zombie rimossa per uid={uid}")

    # Se c'è già un'immagine in attesa, lancia direttamente la generazione
    if uid in pending and pending[uid].get('img'):
        pending[uid]['filter_key'] = fkey
        img = pending[uid]['img']
        if fkey == "artistic_style":
            # Chiedi l'artista prima di generare
            try:
                bot.edit_message_text(
                    "🎨 <b>Stile Artistico</b>\n\nScegli l'artista:",
                    call.message.chat.id, call.message.message_id,
                    parse_mode="HTML", reply_markup=get_artist_keyboard())
            except Exception:
                bot.send_message(call.message.chat.id,
                    "🎨 <b>Stile Artistico</b>\n\nScegli l'artista:",
                    parse_mode="HTML", reply_markup=get_artist_keyboard())
        else:
            logger.info(f"🚀 {fname} → avvio diretto (foto già presente)")
            data = {'filter_key': fkey, 'img': img}
            if fkey == "emoji_art":
                executor.submit(_run_emoji_art, call.message.chat.id, uid, fname, img)
            elif fkey == "lego_galaxy":
                executor.submit(_run_lego_galaxy, call.message.chat.id, uid, fname, img)
            elif fkey == "lego_minifig":
                executor.submit(_run_lego_minifig, call.message.chat.id, uid, fname, img)
            else:
                executor.submit(_run_generation, call.message.chat.id, uid, fname, data)
    else:
        if fkey == "artistic_style":
            try:
                bot.edit_message_text(
                    "🎨 <b>Stile Artistico</b>\n\nScegli l'artista:",
                    call.message.chat.id, call.message.message_id,
                    parse_mode="HTML", reply_markup=get_artist_keyboard())
            except Exception:
                bot.send_message(call.message.chat.id,
                    "🎨 <b>Stile Artistico</b>\n\nScegli l'artista:",
                    parse_mode="HTML", reply_markup=get_artist_keyboard())
        else:
            try:
                bot.edit_message_text(
                    f"✅ Filtro selezionato: <b>{fname}</b>\n\nOra invia la foto da elaborare.",
                    call.message.chat.id, call.message.message_id, parse_mode="HTML")
            except Exception:
                bot.send_message(call.message.chat.id,
                    f"✅ Filtro selezionato: <b>{fname}</b>\n\nOra invia la foto da elaborare.")



# --- 3D STEREO ---
STEREO_PROMPT = (
    "Analyze the outfit description in the subject description above carefully. "
    "Recreate it as a single composite image in 21:9 aspect ratio divided into exactly two equal panels side by side, "
    "designed for cross-eyed stereoscopic 3D viewing.\n\n"
    "LEFT PANEL — right-eye view: recreate the scene with the virtual camera shifted approximately 6cm to the RIGHT. "
    "All lighting, colors, textures and subject identity must remain identical to the original.\n\n"
    "RIGHT PANEL — left-eye view: recreate the exact same scene with the virtual camera shifted approximately 6cm to the LEFT.\n\n"
    "CRITICAL RULES:\n"
    "- Each panel occupies exactly half the total 21:9 width\n"
    "- Foreground elements must show MORE horizontal shift than background elements — this creates natural depth parallax\n"
    "- Background elements shift less; distant elements shift the least\n"
    "- Both panels are photorealistic and visually identical except for the horizontal perspective shift\n"
    "- No color separation, no anaglyph effect, no red/cyan tinting\n"
    "- No visible seam or border between the two panels\n"
    "- Cross-eyed viewing: when the viewer crosses their eyes, the two panels must converge into a single 3D image with natural perceived depth\n"
    "- Maintain all original details: lighting, colors, textures, subject identity, mood\n"
    "- Do NOT add watermarks or text"
)

stereo_waiting = {}  # uid → True se in attesa di foto

@bot.callback_query_handler(func=lambda c: c.data == "stereo_start")
def handle_stereo_start(call):
    uid = call.from_user.id
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    stereo_waiting[uid] = True
    bot.send_message(call.message.chat.id,
        "🎥 <b>3D Stereo</b>\n\n"
        "Invia la foto da convertire in stereoscopico cross-eyed.\n"
        "Funziona meglio con soggetti in primo piano su sfondo distante.")

def _run_stereo(chat_id, uid, username, img_bytes):
    t_start = time.time()
    logger.info(f"   🎥 Stereo richiesto da {username} — generazione immagini rimossa, invio prompt Flow")
    try:
        full_prompt = STEREO_PROMPT
        last_prompt[uid] = full_prompt
        elapsed = round(time.time() - t_start, 1)
        result_markup = types.InlineKeyboardMarkup()
        result_markup.add(types.InlineKeyboardButton("📷 Nuova foto", callback_data="stereo_start"))
        result_markup.add(types.InlineKeyboardButton("🏠 Menu principale", callback_data="stereo_menu"))
        bot.send_message(chat_id, "⚡ <b>Prompt Stereo pronto.</b> Usa Flow per generare.")
        chunks = [full_prompt[i:i+3800] for i in range(0, len(full_prompt), 3800)]
        for idx, chunk in enumerate(chunks):
            header = "" if len(chunks) == 1 else f"<i>({idx+1}/{len(chunks)})</i>\n"
            if idx == len(chunks) - 1:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>", reply_markup=result_markup)
            else:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>")
        logger.info(f"   ✅ Prompt stereo inviato a {username} in {elapsed}s")
    except Exception as e:
        elapsed = round(time.time() - t_start, 1)
        logger.error(f"   ❌ Errore stereo: {e}", exc_info=True)
        bot.send_message(chat_id, f"❌ Errore interno ({elapsed}s):\n<code>{html.escape(str(e))}</code>")

stereo_last_img = {}  # uid → img_bytes per retry

# --- MIRROR OUTFITS ---

def _build_mirror_prompt(outfit_desc):
    """Genera il prompt Mirror con outfit reale dall'analisi immagine."""
    return (
        f"SUBJECT DESCRIPTION — USE THIS EXACTLY:\n"
        f"{outfit_desc}\n\n"
        "CRITICAL — THE SUBJECT MUST ALWAYS BE FULLY CLOTHED IN ALL 4 PANELS. "
        "The outfit described above is MANDATORY and must be reproduced identically. "
        "Generating the subject without clothing or with a different outfit is strictly forbidden.\n\n"
        "IMAGE GENERATION: Generate a single photorealistic image showing a luxury vanity dressing room. "
        "On the makeup table sits a makeup mirror composed of exactly 4 SEPARATE FLAT PANELS arranged side by side, "
        "each panel slightly angled approximately 20 degrees from the next, like a polyptych. "
        "There is NO center panel — all 4 panels are equal in size and clearly distinct. "
        "The subject is seen from behind in the foreground, sitting at the vanity, slightly out of focus. "
        "The subject from behind MUST also be wearing the EXACT SAME OUTFIT described above — "
        "same garments, same colors, same accessories, visible from the back.\n\n"
        "Each of the 4 mirror panels MUST show the subject wearing the EXACT SAME OUTFIT — "
        "identical garments, identical colors, identical accessories, no exceptions, no substitutions.\n\n"
        "The 4 panels differ ONLY in:\n"
        "- Slight angle variation (each panel reflects a marginally different viewpoint)\n"
        "- Subtle expression change (neutral / slight smile / eyes down / eyes up)\n"
        "- Minor head or shoulder pose shift — NEVER a different outfit, NEVER without clothing\n\n"
        "TECHNICAL REQUIREMENTS:\n"
        "- Warm Hollywood vanity lighting with round bulbs framing the mirror\n"
        "- Each panel is a distinct flat mirror — no curved or folding mirrors\n"
        "- Correct mirror reflections with natural perspective for each angle\n"
        "- Subject from behind in foreground is partially visible, out of focus\n"
        "- Each panel reflection is sharp, clearly shows face AND full outfit\n"
        "- Elegant, cinematic atmosphere — luxury dressing room\n"
        "- 8K resolution, 4:3 format\n"
        "- No text, no watermarks"
    )

@bot.callback_query_handler(func=lambda c: c.data in ["stereo_retry", "stereo_menu"])
def handle_stereo_post(call):
    uid = call.from_user.id
    username = call.from_user.username or call.from_user.first_name
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    if call.data == "stereo_menu":
        bot.send_message(call.message.chat.id, "🎨 <b>Scegli una categoria:</b>", reply_markup=cat_keyboard())
        return
    # stereo_retry
    img_bytes = stereo_last_img.get(uid)
    if not img_bytes:
        bot.send_message(call.message.chat.id, "⚠️ Sessione scaduta. Invia di nuovo la foto.")
        stereo_waiting[uid] = True
        return
    bot.send_message(call.message.chat.id, "🔁 Riprovo la stessa foto...\n⏳ Attendi ~30 secondi.")
    executor.submit(_run_stereo, call.message.chat.id, uid, username, img_bytes)

mirror_waiting = {}   # uid → True
mirror_last_img = {}  # uid → img_bytes per retry
mirror_pending_prompt = {}  # uid → prompt string

@bot.callback_query_handler(func=lambda c: c.data == "mirror_confirm")
def handle_mirror_confirm(call):
    uid = call.from_user.id
    username = call.from_user.username or call.from_user.first_name
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    img_bytes = mirror_last_img.get(uid)
    if not img_bytes:
        bot.send_message(call.message.chat.id, "⚠️ Sessione scaduta. Invia di nuovo la foto.")
        return
    bot.send_message(call.message.chat.id, "🪞 <b>Mirror Outfits avviato!</b>\n⏳ Attendi ~30 secondi.")
    executor.submit(_run_mirror, call.message.chat.id, uid, username, img_bytes)

@bot.callback_query_handler(func=lambda c: c.data == "altri_menu")
def handle_altri_menu(call):
    uid = call.from_user.id
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("🎥 3D Stereo", callback_data="stereo_start"))
    markup.add(types.InlineKeyboardButton("🪞 Mirror Outfits", callback_data="mirror_start"))
    for fkey, fdata in filters_by_cat("altri").items():
        markup.add(types.InlineKeyboardButton(fdata["label"], callback_data=f"f_{fkey}"))
    markup.add(types.InlineKeyboardButton("◀️ Indietro", callback_data="back_cats"))
    bot.send_message(call.message.chat.id, "✨ <b>Altri filtri:</b>", reply_markup=markup)

@bot.callback_query_handler(func=lambda c: c.data == "mirror_start")
def handle_mirror_start(call):
    uid = call.from_user.id
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    mirror_waiting[uid] = True
    bot.send_message(call.message.chat.id,
        "🪞 <b>Mirror Outfits</b>\n\n"
        "Invia una foto del soggetto.\n"
        "Genererò 4 riflessi allo specchio della stessa foto, con lo stesso outfit.")

def _run_mirror(chat_id, uid, username, img_bytes):
    t_start = time.time()
    logger.info(f"   🪞 Mirror Outfits | {username} — analisi outfit in corso")
    try:
        # Analisi outfit reale dall'immagine — fino a 2 tentativi
        wait = bot.send_message(chat_id, "🔍 <b>Analizzo l'outfit...</b>")
        outfit_desc, err = analyze_scene(img_bytes, gemini)
        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass

        if not outfit_desc:
            bot.send_message(chat_id, err or "⚠️ Analisi fallita.", parse_mode="HTML")
            logger.error(f"   ❌ Mirror: analyze_scene fallita per {username}")
            return

        prompt = _build_mirror_prompt(outfit_desc)
        last_prompt[uid] = prompt
        elapsed = round(time.time() - t_start, 1)
        result_markup = types.InlineKeyboardMarkup()
        result_markup.add(types.InlineKeyboardButton("📷 Nuova foto", callback_data="mirror_start"))
        result_markup.add(types.InlineKeyboardButton("🏠 Menu principale", callback_data="mirror_menu"))
        bot.send_message(chat_id, "⚡ <b>Prompt Mirror pronto.</b> Usa Flow per generare.")
        chunks = [prompt[i:i+3800] for i in range(0, len(prompt), 3800)]
        for idx, chunk in enumerate(chunks):
            header = "" if len(chunks) == 1 else f"<i>({idx+1}/{len(chunks)})</i>\n"
            if idx == len(chunks) - 1:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>", reply_markup=result_markup)
            else:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>")
        logger.info(f"   ✅ Prompt mirror inviato a {username} in {elapsed}s")
    except Exception as e:
        elapsed = round(time.time() - t_start, 1)
        logger.error(f"   ❌ Errore mirror: {e}", exc_info=True)
        bot.send_message(chat_id, f"❌ Errore interno ({elapsed}s):\n<code>{html.escape(str(e))}</code>")

@bot.callback_query_handler(func=lambda c: c.data in ["mirror_retry", "mirror_menu"])
def handle_mirror_post(call):
    uid = call.from_user.id
    username = call.from_user.username or call.from_user.first_name
    try:
        bot.answer_callback_query(call.id)
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass
    if call.data == "mirror_menu":
        bot.send_message(call.message.chat.id, "🎨 <b>Scegli una categoria:</b>", reply_markup=cat_keyboard())
        return
    img_bytes = mirror_last_img.get(uid)
    if not img_bytes:
        bot.send_message(call.message.chat.id, "⚠️ Sessione scaduta. Invia di nuovo la foto.")
        mirror_waiting[uid] = True
        return
    bot.send_message(call.message.chat.id, "🔁 Riprovo con nuovi outfit...\n⏳ Attendi ~30 secondi.")
    executor.submit(_run_mirror, call.message.chat.id, uid, username, img_bytes)
@bot.callback_query_handler(func=lambda c: c.data.startswith("mosaic_build_"))
def handle_mosaic_build(call):
    uid = call.from_user.id
    bot.answer_callback_query(call.id)
    if uid not in mosaic_collecting:
        bot.send_message(call.message.chat.id, "⚠️ Sessione mosaic scaduta. Usa /mosaic per ricominciare.")
        return
    executor.submit(_finalize_mosaic, uid, call.message.chat.id)

# --- /caption standalone ---
@bot.message_handler(commands=['caption'])
def cmd_caption(m):
    uid = m.from_user.id
    img = last_img.get(uid)
    if not img:
        bot.send_message(m.chat.id, "📸 Invia prima una foto, poi usa /caption.")
        return
    wait = bot.send_message(m.chat.id, "✍️ Genero caption...")
    result, err = caption.from_image(img)
    try: bot.delete_message(m.chat.id, wait.message_id)
    except Exception: pass
    if result:
        bot.send_message(m.chat.id, result)
    else:
        bot.send_message(m.chat.id, err or "⚠️ Caption non disponibile.", parse_mode="HTML")

# --- MOSAIC ---
def _start_mosaic_session(uid, chat_id, username):
    if uid in mosaic_collecting and mosaic_collecting[uid].get('timer'):
        mosaic_collecting[uid]['timer'].cancel()
    mosaic_collecting[uid] = {'photos': [], 'timer': None, 'aspect': None}
    logger.info(f"🖼️ Mosaic avviato da {username} (id={uid})")
    bot.send_message(chat_id,
        "🖼️ <b>Modalità Mosaic attiva!</b>\n\n"
        "Invia da <b>2 a 9 foto</b> e premi <b>✅ Procedi</b> quando vuoi.\n"
        "⏱️ La sessione scade automaticamente dopo 5 minuti di inattività."
    )

@bot.message_handler(commands=['mosaic'])
def cmd_mosaic(m):
    _start_mosaic_session(m.from_user.id, m.chat.id, m.from_user.username or m.from_user.first_name)

@bot.callback_query_handler(func=lambda c: c.data == "mosaic_restart")
def cb_mosaic_restart(call):
    uid = call.from_user.id
    if not is_allowed(uid): return
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    try: bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception: pass
    _start_mosaic_session(uid, call.message.chat.id,
                          call.from_user.username or call.from_user.first_name)


@bot.callback_query_handler(func=lambda c: c.data == "mosaic_start")
def cb_mosaic_start(call):
    uid = call.from_user.id
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    try: bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception: pass
    _start_mosaic_session(uid, call.message.chat.id, call.from_user.username or call.from_user.first_name)

@bot.callback_query_handler(func=lambda c: c.data == "artist_cat_noop")
def cb_artist_noop(call):
    try: bot.answer_callback_query(call.id)
    except Exception: pass

@bot.callback_query_handler(func=lambda c: c.data == "back_artistic")
def cb_back_artistic(call):
    """Torna al menu categorie artisti."""
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    try:
        bot.edit_message_text("🎨 <b>Stile Artistico</b> — scegli la categoria:",
            call.message.chat.id, call.message.message_id,
            reply_markup=get_artist_cat_keyboard(), parse_mode="HTML")
    except Exception:
        bot.send_message(call.message.chat.id, "🎨 <b>Stile Artistico</b> — scegli la categoria:",
            reply_markup=get_artist_cat_keyboard(), parse_mode="HTML")

@bot.callback_query_handler(func=lambda c: c.data.startswith("artcat_"))
def cb_artcat(call):
    """Mostra lista artisti per una categoria."""
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    cat_key = call.data.replace("artcat_", "", 1)
    if cat_key not in ARTIST_CAT_MAP:
        return
    cat_label, _ = ARTIST_CAT_MAP[cat_key]
    try:
        bot.edit_message_text(f"🎨 <b>{cat_label}</b> — scegli l'artista:",
            call.message.chat.id, call.message.message_id,
            reply_markup=get_artist_list_keyboard(cat_key), parse_mode="HTML")
    except Exception:
        bot.send_message(call.message.chat.id, f"🎨 <b>{cat_label}</b> — scegli l'artista:",
            reply_markup=get_artist_list_keyboard(cat_key), parse_mode="HTML")

@bot.callback_query_handler(func=lambda c: c.data.startswith("artcat_"))
def cb_artist_cat(call):
    """Mostra lista artisti di una categoria."""
    uid = call.from_user.id
    if not is_allowed(uid): return
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    cat_label = call.data.replace("artcat_", "", 1)
    try:
        bot.edit_message_text(f"🎨 <b>{cat_label}</b> — scegli l'artista:",
            call.message.chat.id, call.message.message_id,
            reply_markup=artist_list_keyboard(cat_label), parse_mode="HTML")
    except Exception:
        bot.send_message(call.message.chat.id, f"🎨 <b>{cat_label}</b> — scegli l'artista:",
            reply_markup=artist_list_keyboard(cat_label), parse_mode="HTML")

@bot.callback_query_handler(func=lambda c: c.data.startswith("artist_"))
def cb_artist_choice(call):
    uid = call.from_user.id
    if not is_allowed(uid):
        return
    artist_key = call.data.replace("artist_", "", 1)
    if artist_key not in ARTISTIC_STYLE_PROMPTS:
        try: bot.answer_callback_query(call.id, "Artista non trovato.")
        except Exception: pass
        return
    user_artist[uid] = artist_key
    user_filter[uid] = "artistic_style"
    try: bot.answer_callback_query(call.id, f"✅ {artist_key}")
    except Exception: pass
    logger.info(f"🎨 {call.from_user.username or call.from_user.first_name} (id={uid}) → artista: {artist_key}")

    # Se c'è già un'immagine in attesa, lancia direttamente la generazione
    if uid in pending and pending[uid].get('img'):
        img = pending[uid]['img']
        data = {'filter_key': 'artistic_style', 'img': img, 'artist_key': artist_key}
        try:
            bot.edit_message_text(
                f"🚀 <b>Generazione avviata!</b>\n🎨 Artista: <b>{artist_key}</b>\n⏳ Attendi ~20–35 secondi.",
                call.message.chat.id, call.message.message_id, parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id,
                f"🚀 <b>Generazione avviata!</b>\n🎨 Artista: <b>{artist_key}</b>\n⏳ Attendi ~20–35 secondi.")
        executor.submit(_run_generation, call.message.chat.id, uid, artist_key, data)
    else:
        try:
            bot.edit_message_text(
                f"🎨 Artista selezionato: <b>{artist_key}</b>\n\nOra invia la foto da elaborare.",
                call.message.chat.id, call.message.message_id, parse_mode="HTML")
        except Exception:
            bot.send_message(call.message.chat.id,
                f"🎨 Artista selezionato: <b>{artist_key}</b>\n\nOra invia la foto da elaborare.")

@bot.message_handler(commands=['done'])
def cmd_done(m):
    uid = m.from_user.id
    if uid not in mosaic_collecting:
        bot.send_message(m.chat.id, "⚠️ Nessuna sessione mosaic attiva. Usa /mosaic per iniziare.")
        return
    executor.submit(_finalize_mosaic, uid, m.chat.id)

def _detect_aspect_label(w, h):
    ratio = w / h
    if ratio > 1.6:
        return "16:9"
    elif ratio > 1.1:
        return "4:3"
    elif ratio > 0.85:
        return "1:1"
    elif ratio > 0.55:
        return "3:4"
    else:
        return "9:16"

def _finalize_mosaic(uid, chat_id):
    data = mosaic_collecting.pop(uid, None)
    if not data:
        return
    if data.get('timer'):
        data['timer'].cancel()

    photos = data['photos']
    n = len(photos)
    aspect_label = data.get('aspect', '1:1')

    if n < 2:
        bot.send_message(chat_id, "⚠️ Servono almeno 2 foto per il collage.")
        return

    # Layout per N=3 dipende dall'aspect ratio
    IS_VERTICAL = aspect_label in ("9:16", "3:4")
    IS_HORIZONTAL = aspect_label in ("16:9", "4:3")

    # LAYOUTS: lista di righe, ogni riga = numero di colonne
    if n == 3:
        layout = None  # gestito separatamente: 1 grande (2/3) + 2 piccole (1/3)
    elif n == 2:
        if IS_HORIZONTAL:
            layout = [(1,), (1,)]  # 2 impilate → verticale
        else:
            layout = [(2,)]        # 2 affiancate → orizzontale
    else:
        LAYOUTS = {
            4: None,  # gestito separatamente: 1 grande + 3 piccole
            5: [(2,), (3,)],
            6: [(3,), (3,)],
            7: [(3,), (4,)],
            8: [(4,), (4,)],
            9: [(3,), (3,), (3,)],
        }
        layout = LAYOUTS[n]

    if n == 3:
        grid_label = "1+2"
    elif n == 4:
        grid_label = "1+3"
    elif n == 6:
        grid_label = "2×3"
    elif n == 8:
        grid_label = "2×4"
    elif n == 9:
        grid_label = "3×3"
    else:
        grid_label = "+".join(str(r[0]) for r in layout)

    bot.send_message(chat_id, f"🖼️ Assemblando collage {grid_label} ({aspect_label})...")

    def get_border_color(img):
        """Colore medio dei 4 bordi — campiona su resize 50px per velocità."""
        small = img.resize((50, 50), Image.BOX)
        w, h = small.size
        pixels = []
        for x in range(w):
            pixels.append(small.getpixel((x, 0)))
            pixels.append(small.getpixel((x, h - 1)))
        for y in range(h):
            pixels.append(small.getpixel((0, y)))
            pixels.append(small.getpixel((w - 1, y)))
        r = sum(p[0] for p in pixels) // len(pixels)
        g = sum(p[1] for p in pixels) // len(pixels)
        b = sum(p[2] for p in pixels) // len(pixels)
        return (r, g, b)

    def fit_tile(img, tw, th):
        """Scala mantenendo aspect ratio, bande con colore bordi."""
        ratio = min(tw / img.width, th / img.height)
        new_w = int(img.width * ratio)
        new_h = int(img.height * ratio)
        img_resized = img.resize((new_w, new_h), Image.LANCZOS)
        bg_color = get_border_color(img)
        tile = Image.new("RGB", (tw, th), bg_color)
        tile.paste(img_resized, ((tw - new_w) // 2, (th - new_h) // 2))
        return tile

    def build():
        try:
            images = [Image.open(io.BytesIO(b)).convert("RGB") for b in photos]

            ref = images[0]
            orig_w, orig_h = ref.width, ref.height
            photo_ratio = orig_h / orig_w  # es. 2496/1664 ≈ 1.5

            # Target grid ~4Mpx — calcolo dimensioni da target, non dalle foto originali
            TARGET_PX = 4_000_000

            # Casi speciali n=3 / n=4: 1 grande + colonna/riga di piccole con rapporto geometrico esatto
            if n in (3, 4):
                k = n - 1
                # TARGET = (k * small_w) * ((k + 1) * small_h) = k*(k+1)*small_w^2*photo_ratio
                small_w = max(1, int((TARGET_PX / (k * (k + 1) * photo_ratio)) ** 0.5))
                small_h = max(1, int(small_w * photo_ratio))
                bg = get_border_color(images[0])

                if IS_HORIZONTAL:
                    # 1 grande sopra, k piccole sotto; ogni piccola ha larghezza = 1/k della grande
                    total_w = k * small_w
                    big_h = k * small_h
                    total_h = big_h + small_h
                    grid = Image.new("RGB", (total_w, total_h), bg)
                    grid.paste(fit_tile(images[0], total_w, big_h), (0, 0))
                    for i, img in enumerate(images[1:n]):
                        grid.paste(fit_tile(img, small_w, small_h), (i * small_w, big_h))
                else:
                    # 1 grande a sinistra, k piccole a destra; ogni piccola ha altezza = 1/k della grande
                    big_w = k * small_w
                    total_w = big_w + small_w
                    total_h = k * small_h
                    grid = Image.new("RGB", (total_w, total_h), bg)
                    grid.paste(fit_tile(images[0], big_w, total_h), (0, 0))
                    for i, img in enumerate(images[1:n]):
                        grid.paste(fit_tile(img, small_w, small_h), (big_w, i * small_h))

            else:
                # Calcola tile_w e tile_h dal TARGET_PX per n≠4
                # Per layout a N colonne: total_px = n_cols * tile_w * tile_w * photo_ratio * n_rows
                first_row_cols = layout[0][0]
                n_rows = len(layout)
                tile_w = int((TARGET_PX / (first_row_cols * n_rows * photo_ratio)) ** 0.5)
                tile_h = int(tile_w * photo_ratio)

                row_images = []
                idx = 0
                total_w = None
                for row_def in layout:
                    n_cols = row_def[0]
                    row_imgs = images[idx:idx + n_cols]
                    idx += n_cols

                    if total_w is None:
                        cell_w = tile_w
                        total_w = cell_w * n_cols
                    else:
                        cell_w = total_w // n_cols

                    cell_h = tile_h

                    tiles = [fit_tile(img, cell_w, cell_h) for img in row_imgs]
                    row_img = Image.new("RGB", (total_w, cell_h), (0, 0, 0))
                    for i, tile in enumerate(tiles):
                        row_img.paste(tile, (i * cell_w, 0))
                    row_images.append(row_img)

                total_h = sum(r.height for r in row_images)
                grid = Image.new("RGB", (total_w, total_h), (0, 0, 0))
                y = 0
                for row_img in row_images:
                    grid.paste(row_img, (0, y))
                    y += row_img.height

            out = io.BytesIO()
            grid.save(out, format="JPEG", quality=92)
            out.seek(0)

            bot.send_document(
                chat_id,
                out,
                visible_file_name=f"mosaic_{n}foto.jpg",
                caption=f"✅ Collage {grid_label} ({aspect_label}) — {n} foto"
            )
            logger.info(f"   ✅ Mosaic {grid_label} inviato a uid={uid}")
            mosaic_markup = types.InlineKeyboardMarkup()
            mosaic_markup.row(
                types.InlineKeyboardButton("🖼️ Nuovo mosaic", callback_data="mosaic_restart"),
                types.InlineKeyboardButton("🔄 Cambia filtro", callback_data="post_newfilter")
            )
            mosaic_markup.row(
                types.InlineKeyboardButton("🆕 Nuova foto e nuovo filtro", callback_data="post_newboth")
            )
            bot.send_message(chat_id, "Cosa vuoi fare adesso?", reply_markup=mosaic_markup)
        except Exception as e:
            logger.error(f"   ❌ Errore assemblaggio mosaic: {e}", exc_info=True)
            bot.send_message(chat_id, f"❌ Errore nell'assemblaggio:\n<code>{html.escape(str(e))}</code>")

    executor.submit(build)

@bot.message_handler(content_types=['photo'])
def handle_photo(m):
    uid = m.from_user.id
    if not is_allowed(uid):
        logger.warning(f"🚫 Foto non autorizzata: uid={uid} username={m.from_user.username}")
        return
    username = m.from_user.username or m.from_user.first_name

    try:
        file_info = bot.get_file(m.photo[-1].file_id)
        img_data = bot.download_file(file_info.file_path)
        logger.info(f"🖼️ Foto da {username} (id={uid}), {len(img_data)} bytes")
    except Exception as e:
        logger.error(f"❌ Errore download foto da {username}: {e}")
        bot.reply_to(m, "❌ Errore nel scaricare la foto. Riprova.")
        return

    # ── MODALITÀ STEREO ───────────────────────────────────────────────────────
    if stereo_waiting.pop(uid, False):
        bot.send_message(m.chat.id,
            "🎥 <b>Elaborazione stereo avviata!</b>\n"
            "⏳ Attendi ~30 secondi.")
        executor.submit(_run_stereo, m.chat.id, uid, username, img_data)
        return

    # ── MODALITÀ MIRROR ───────────────────────────────────────────────────────
    if mirror_waiting.pop(uid, False):
        mirror_last_img[uid] = img_data
        wait = bot.send_message(m.chat.id, "🔍 <b>Analizzo l'outfit...</b>")
        outfit_desc, err = analyze_scene(img_data, gemini)
        try: bot.delete_message(m.chat.id, wait.message_id)
        except Exception: pass
        if not outfit_desc:
            bot.send_message(m.chat.id, err or "⚠️ Analisi fallita.", parse_mode="HTML")
            return
        prompt = _build_mirror_prompt(outfit_desc)
        # Invia prompt in anteprima con pulsante conferma
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🚀 Genera", callback_data="mirror_confirm"))
        markup.add(types.InlineKeyboardButton("❌ Annulla", callback_data="mirror_menu"))
        chunks = [prompt[i:i+3800] for i in range(0, len(prompt), 3800)]
        for idx, chunk in enumerate(chunks):
            header = "🪞 <b>Mirror Outfits — Prompt generato:</b>\n\n" if idx == 0 else f"<i>(continua {idx+1}/{len(chunks)})</i>\n\n"
            bot.send_message(m.chat.id, f"{header}<code>{html.escape(chunk)}</code>",
                reply_markup=markup if idx == len(chunks)-1 else None)
        # Salva prompt per uso nel callback
        mirror_pending_prompt[uid] = prompt
        return

    # ── MODALITÀ MOSAIC ───────────────────────────────────────────────────────
    if uid in mosaic_collecting:
        session = mosaic_collecting[uid]
        if session.get('timer'):
            session['timer'].cancel()
        session['photos'].append(img_data)
        n = len(session['photos'])
        logger.info(f"   🖼️ Mosaic raccolta: {n} foto per uid={uid}")

        # Rileva aspect ratio dalla prima foto
        if n == 1:
            try:
                first_img = Image.open(io.BytesIO(img_data))
                w, h = first_img.size
                session['aspect'] = _detect_aspect_label(w, h)
                logger.info(f"   📐 Aspect rilevato: {session['aspect']} ({w}x{h})")
            except Exception:
                session['aspect'] = '?'

        aspect_label = session.get('aspect', '?')

        if n >= 9:
            # Massimo raggiunto — forza assemblaggio
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton(
                f"✅ Procedi ({n} foto)",
                callback_data=f"mosaic_build_{uid}"
            ))
            bot.reply_to(m,
                f"📷 <b>{n} foto</b> — massimo raggiunto!\n"
                f"Formato: <b>{aspect_label}</b>",
                reply_markup=markup)
        elif n >= 2:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton(
                f"✅ Procedi ({n} foto)",
                callback_data=f"mosaic_build_{uid}"
            ))
            bot.reply_to(m,
                f"📷 <b>{n} foto</b> ricevute — puoi procedere o inviarne altre (max 9).\n"
                f"Formato: <b>{aspect_label}</b>",
                reply_markup=markup)
        else:
            # n == 1
            bot.reply_to(m, f"📷 Prima foto ricevuta. Invia almeno un'altra per il collage.")

        # Timer 5 minuti safety net
        t = threading.Timer(300, _finalize_mosaic, args=[uid, m.chat.id])
        t.daemon = True
        t.start()
        session['timer'] = t
        return
    # FIX 2.0.0 (bug #9): rimosso blocco "cleanup sessione zombie" — era un secondo
    # `if uid in mosaic_collecting` identico al precedente, quindi mai raggiungibile
    # (il branch sopra fa sempre return). Verificato che non serve comunque: ogni
    # uscita da una sessione mosaic passa già da _finalize_mosaic() o
    # _start_mosaic_session(), che ripuliscono mosaic_collecting[uid] correttamente.

    fkey = user_filter[uid]
    pending[uid] = {'img': img_data, 'filter_key': fkey}
    last_img[uid] = img_data

    if not fkey:
        bot.reply_to(m, "📷 Foto ricevuta!\n\n🎨 <b>Scegli una categoria:</b>", reply_markup=cat_keyboard())
    elif fkey == "artistic_style" and uid not in user_artist:
        # Artista non ancora scelto — mostra menu categorie artisti
        bot.reply_to(m, "🎨 <b>Stile Artistico</b>\n\nScegli la categoria:",
            parse_mode="HTML", reply_markup=artist_cat_keyboard())
    else:
        fname = FILTERS[fkey]["label"]
        logger.info(f"🚀 {username} (id={uid}) → avvio diretto | filtro: {fkey}")
        artist_key = user_artist.get(uid) if fkey == "artistic_style" else None
        data = {'filter_key': fkey, 'img': img_data}
        if artist_key:
            data['artist_key'] = artist_key
        if fkey == "emoji_art":
            executor.submit(_run_emoji_art, m.chat.id, uid, username, img_data)
        elif fkey == "lego_galaxy":
            executor.submit(_run_lego_galaxy, m.chat.id, uid, username, img_data)
        elif fkey == "lego_minifig":
            executor.submit(_run_lego_minifig, m.chat.id, uid, username, img_data)
        else:
            executor.submit(_run_generation, m.chat.id, uid, username, data)



FULLBODY_3X3_BASE = (
    "Using the subject description above as the identity and style reference, create a full body editorial grid "
    "with EXACTLY 3 columns and 3 rows — 9 panels total, square or landscape orientation. "
    "The grid MUST be 3 wide by 3 tall. Strictly 3x3, no other layout. "
    "IDENTITY LOCK: Preserve the subject's identity — same face, facial structure, skin tone, hair, and use the EXACT outfit described in the subject description above — "
    "outfit, shoes, accessories, jewelry, and any distinctive elements visible in the image. "
    "Replicate the outfit completely and consistently across every single panel. "
    "FRAMING: Each panel shows the FULL BODY from head to toe — outfit fully visible in every panel. "
    "Wide enough framing to include the complete silhouette, shoes included. "
    "SETTING: Use the background environment and lighting mood described above. "
    "Natural realistic skin tones. Medium to high contrast. Thin gutters between panels. High panel consistency. "
    "85mm lens look. Cinematic editorial quality. "
    "9 full body poses (randomly selected, in this order): "
    "{poses} "
    "Ultra high resolution photorealistic editorial fashion contact sheet. "
)



def build_y2k_pop_collage_prompt(outfit_desc=""):
    """Pesca 5 pose casuali dalla Y2K_POSE_POOL e costruisce il prompt collage."""
    poses = random.sample(Y2K_POSE_POOL, 5)
    panels = "\n".join(
        f"Panel {i+1}: {pose}."
        for i, pose in enumerate(poses)
    )
    return (
        "Generate an ultra-realistic playful fashion editorial collage of the subject and outfit "
        "described above. "
        "Style: bold Y2K pop-star aesthetic, luxury playful campaign vibe inspired by early 2000s "
        "celebrity magazine editorials, high-fashion meets cartoon-core, fun exaggerated energy. "
        "Setting: seamless white studio backdrop, bright diffused studio lighting, soft shadow "
        "gradients, glossy fashion magazine finish, ultra-clean editorial composition. "
        "Layout: 5 vertical panels arranged side by side in a single wide collage, "
        "thin white gutters between panels. "
        "All 5 panels show the exact same subject with the exact same outfit described above — "
        "no outfit changes, no color shifts, full consistency across all panels. "
        f"{panels}\n"
        "Lighting consistent across all panels: bright even studio softboxes, subtle rim light, "
        "clean white floor reflection, ultra-polished glossy magazine finish. "
        "Ultra high resolution, photorealistic, 8K, no text, no watermark, no blur."
    )


def resolve_prompt(fkey):
    """Risolve il prompt del filtro. Se è callable (lambda), lo esegue e costruisce il testo finale."""
    raw = FILTERS[fkey]["prompt"]
    if fkey == "artistic_style":
        prompt, _ = build_artistic_style_prompt()
        return prompt
    if fkey == "y2k_pop_collage":
        return build_y2k_pop_collage_prompt()
    if callable(raw):
        if fkey == "fullbody_3x3":
            poses = raw()
            pose_str = ", ".join(f"{i+1}-{p}" for i, p in enumerate(poses))
            return FULLBODY_3X3_BASE.format(poses=pose_str)
        return raw()
    return raw

# Salva lo stile estratto per mostrarlo nella preview
pending_artistic_style = {}  # uid → style label estratto


# --- GENERAZIONE ---
# ─── EMOJI ART ────────────────────────────────────────────────────────────────
# ─── LEGO MOSAIC ──────────────────────────────────────────────────────────────
# Palette LEGO ufficiale — nome, R, G, B
LEGO_PALETTE = [
    # Bianchi / grigi / neri
    ("White",               242, 243, 242),
    ("Light Bluish Gray",   171, 173, 172),
    ("Medium Stone Grey",   163, 162, 165),
    ("Dark Bluish Gray",    108, 110, 104),
    ("Dark Stone Grey",      74,  73,  74),
    ("Black",                27,  42,  52),
    # Gialli
    ("Yellow",              247, 209,  23),
    ("Bright Yellow",       255, 205,   0),
    ("Warm Gold",           170, 127,  46),
    ("Tan",                 222, 198, 156),
    ("Light Nougat",        249, 193, 152),
    ("Nougat",              204, 142,  87),
    ("Medium Nougat",       170, 125,  85),
    ("Dark Nougat",         170,  85,  10),
    ("Reddish Brown",       105,  64,  40),
    ("Dark Brown",           53,  33,   0),
    # Rossi / arancio / rosa
    ("Red",                 179,   0,   6),
    ("Bright Red",          200,   0,   8),
    ("Dark Red",            114,  14,  15),
    ("Orange",              213,  99,  22),
    ("Bright Orange",       255, 126,  20),
    ("Dark Orange",         157,  79,  12),
    ("Coral",               255, 109,  88),
    ("Salmon",              255, 144, 119),
    ("Light Pink",          255, 173, 185),
    ("Pink",                255, 105, 143),
    ("Magenta",             144,  31, 118),
    ("Dark Pink",           201,   0,  70),
    # Verdi
    ("Bright Green",          0, 185,   0),
    ("Green",                 0, 146,  74),
    ("Dark Green",            0, 100,  46),
    ("Lime Green",           125, 189,   0),
    ("Medium Green",         172, 210, 152),
    ("Olive Green",           75,  84,  32),
    ("Sand Green",           120, 154, 136),
    # Blu
    ("Blue",                  0,  85, 191),
    ("Bright Blue",            0, 114, 206),
    ("Dark Blue",              0,  32, 160),
    ("Medium Blue",           97, 175, 255),
    ("Sky Blue",             135, 205, 234),
    ("Dark Azure",             0, 131, 203),
    ("Medium Azure",           0, 168, 222),
    ("Sand Blue",              95, 124, 148),
    # Viola
    ("Purple",                81,   0, 125),
    ("Medium Lilac",          52,  43, 117),
    ("Lavender",             160, 150, 220),
    ("Medium Lavender",      169, 127, 200),
    ("Sand Purple",          135, 124, 144),
]
_LEGO_PALETTE_BUILT = [(r, g, b, name) for name, r, g, b in LEGO_PALETTE]

def _closest_lego(r, g, b):
    """Trova il colore LEGO più vicino con distanza euclidea pesata (percezione umana)."""
    best, best_dist = "White", float("inf")
    for pr, pg, pb, name in _LEGO_PALETTE_BUILT:
        d = (2*(r-pr)**2 + 4*(g-pg)**2 + 3*(b-pb)**2) ** 0.5
        if d < best_dist:
            best_dist, best = d, name
    return best

# Stato per lista mattoncini
lego_waiting_list = {}   # uid → {'counts': dict, 'cols': int, 'rows': int}

def _draw_stud(draw, x, y, size, rgb):
    """Disegna un mattoncino 1×1 con stud e effetto 3D sottile."""
    shadow    = tuple(max(0,   c - 45) for c in rgb)
    highlight = tuple(min(255, c + 45) for c in rgb)
    mid_hi    = tuple(min(255, c + 20) for c in rgb)

    # Corpo mattoncino
    draw.rectangle([x, y, x+size-1, y+size-1], fill=rgb)
    # Highlight top-left
    draw.line([x, y, x+size-1, y], fill=highlight, width=2)
    draw.line([x, y, x, y+size-1], fill=highlight, width=2)
    # Ombra bottom-right
    draw.line([x, y+size-1, x+size-1, y+size-1], fill=shadow, width=2)
    draw.line([x+size-1, y, x+size-1, y+size-1], fill=shadow, width=2)

    # Stud rialzato: cerchio con highlight sopra e ombra sotto
    pad = int(size * 0.22)
    cx0, cy0, cx1, cy1 = x+pad, y+pad, x+size-pad, y+size-pad
    draw.ellipse([cx0, cy0, cx1, cy1], fill=mid_hi, outline=shadow)
    # Piccolo arco highlight in alto a sinistra dello stud
    inner = int((cx1-cx0) * 0.25)
    draw.arc([cx0, cy0, cx0+inner*2, cy0+inner*2], start=200, end=320, fill=highlight, width=max(1, size//14))

# Pool location per LEGO Minifig — attinge dallo stesso spirito di Surprise
_MINIFIG_LOCATIONS = [
    "Apollo lunar landing site, Moon surface, grey cratered regolith, NASA Lunar Module Eagle with gold foil legs, astronaut in white EVA suit with mirrored visor, Earth rising above lunar horizon, black star-filled space, no atmosphere, harsh directional sunlight casting sharp shadows, desolate grandeur, photorealistic NASA archive aesthetic",
    "Mars surface, Jezero Crater, rusty red iron-oxide dust plain, NASA Perseverance rover beside subject, panoramic Martian horizon, pink-ochre dusty sky, low sun angle, distant mesa formations, tyre tracks in red soil, scientific equipment visible, photorealistic NASA Mars mission aesthetic, epic isolation",
    "2001 A Space Odyssey aesthetic, floating in Jupiter orbit, massive perfectly black rectangular Monolith hovering in space, Jupiter's swirling Great Red Spot storm below, Stanley Kubrick ultra-precise composition, blue dawn light of star alignment, Ligeti Atmosphères atmosphere, primordial cosmic awe, photorealistic hard sci-fi",
    "Saturn's rings close-up perspective, subject skating on flat ice-particle ring surface, Saturn's golden banded atmosphere filling half the sky above, other ring sheets visible in layers, distant sun, deep space beyond, ring particles ranging from dust to boulders visible around, impossible dreamlike scale, photorealistic space photography aesthetic",
    "Alien planet surface, bioluminescent purple and teal vegetation, double moons in sky, crystalline rock formations, two or three friendly alien beings of original non-human design standing near subject, alien architecture in background, impossible colors, James Cameron Avatar-scale world-building, cinematic photorealistic alien landscape",
    "Inside spacecraft cockpit traversing wormhole, Contact film 1997 Zemeckis aesthetic, swirling iridescent tunnel of spacetime with concentric light rings stretching to infinity, stars smearing into streaks, pod chair visible, interior bathed in warm amber and cold blue light alternating, Jodie Foster wide-eyed wonder expression reference, photorealistic cinematic",
    "Back to the Future DeLorean DMC-12 time machine, stainless steel gull-wing doors open, flux capacitor glowing inside, Mr. Fusion reactor on back, lightning storm at Twin Pines Mall parking lot, 88mph fire trails on asphalt, Robert Zemeckis 1985 cinematic aesthetic, neon rain, photorealistic film still",
    "Tim Burton Gotham City 1989 aesthetic, original Batmobile elongated jet-turbine car on rain-slicked gothic street, art deco gargoyle skyscrapers above, searchlight bat-signal in purple night sky, subject beside open cockpit, dramatic chiaroscuro street lighting, Danny Elfman atmosphere, photorealistic film noir",
    "Mos Eisley Cantina, Tatooine, Star Wars A New Hope 1977 aesthetic, dimly lit alien bar interior, Figrin D'an and the Modal Nodes playing cantina band music, diverse alien species at rounded booths, Han Solo's booth in background, warm amber practical lighting, bar stools and drinks, John Williams cantina music visual energy, photorealistic film still",
    "Forest moon of Endor, Star Wars Return of the Jedi, ancient redwood-scale trees, Ewok village rope bridges and wooden platforms high in canopy, Ewoks in background, dappled forest light, Imperial AT-AT walker silhouette through trees, warm golden forest light, John Williams adventure score energy, photorealistic film still",
    "Tatooine twin sunset, Star Wars, red rock canyon walls, Lars homestead dome visible on flat desert horizon, twin suns setting in amber sky, binary sunset Luke Skywalker composition reference, desert wind dust, remote alien frontier, John Williams Force theme visual atmosphere, photorealistic film still",
    "USS Enterprise bridge, Star Trek Original Series 1966 aesthetic, circular command centre, Kirk's captain chair, colourful blinking control panels, Spock at science station hooded viewer, uhura communications earpiece, primary colour crew uniforms, viewscreen with stars, Alexander Courage fanfare visual energy, photorealistic 1960s TV production design",
    "Planet Vulcan surface, Star Trek Original Series aesthetic, red-orange desert landscape, twin moons, angular modernist Vulcan architecture in stone and metal, Mount Seleya in background, ceremonial grounds, logic and discipline atmosphere, ancient alien civilisation grandeur, photorealistic 1960s science fiction aesthetic updated to cinematic quality",
    "Star Trek Original Series alien planet city set, primary colored geometric architecture, M-class planet with orange sky, classic series transporter beam effect, communicator prop visible, Tricorder scanning gesture, James T. Kirk swaggering diplomacy energy, bold primary color palette, photorealistic 1960s science fiction aesthetic",
    "Interstellar tesseract interior, Christopher Nolan 2014, infinite bookshelf corridor bending into five dimensions, Cooper falling through timeframes, hands reaching through shelves, gravitational anomaly sand patterns, warm dust-light filtering through dimensional cracks, Hans Zimmer pipe organ score visual energy, IMAX cinematic quality, photorealistic",
    "Miller's water planet, Interstellar, Gargantua black hole looming impossibly large on horizon warping light behind it, vast shallow ocean knee-deep, enormous tidal wave wall kilometres high on horizon, wrecked Ranger spacecraft partly submerged, alien blue-white light, Hans Zimmer Cornfield Chase visual atmosphere, photorealistic IMAX cinematic",
    "X-Files aesthetic, rural Oregon cornfield at night, two FBI agents in dark suits with flashlights — male with loosened tie, female with red hair — looking up at hovering flying saucer craft with blinding white light beam, government black SUVs in background, Mark Snow paranoid score atmosphere, photorealistic 1990s TV film noir",
    "Psychedelic surrealist dreamscape, Salvador Dali meets Yellow Submarine Beatles animation, melting clocks and impossible architecture, kaleidoscopic fractal sky in magenta violet gold, giant flowers as tall as buildings, checkered floor dissolving into clouds, M.C. Escher impossible staircases, Rick Griffin psychedelic poster palette, everything hyper-detailed and iridescent",
    "The Shire, Lord of the Rings Peter Jackson trilogy aesthetic, rolling green hills of Hobbiton, round hobbit hole doors with circular windows, manicured gardens with giant vegetables, Party Tree in background, warm afternoon golden light, Bag End chimney smoke, Howard Shore pastoral score visual energy, photorealistic New Zealand landscape",
    "Mount Doom, Mordor, Lord of the Rings Peter Jackson, active volcanic crater edge with molten lava rivers, Barad-dûr tower with lidless Eye of Sauron in dark sky, ash falling like snow, red hellish atmosphere, extreme heat haze, Howard Shore dramatic score visual energy, photorealistic WETA Workshop cinematic quality",
    "Rivendell, Lord of the Rings, elven architecture of waterfalls and elegant stone bridges, autumn forest canopy in gold and amber, Hall of Fire interior with carved pillars, Alan Lee concept art aesthetic, Howard Shore elven theme visual atmosphere, long shadows and mist over the valley, photorealistic Peter Jackson trilogy quality",
    "Great Hall of Hogwarts, Harry Potter, enchanted ceiling showing night sky with floating candles, four house tables, owls delivering post, golden goblets and magical feast, John Williams Hedwig's Theme visual energy, stone Gothic arches, torchlight on ancient banners, photorealistic Warner Bros film aesthetic, back-to-school wonder",
    "Jurassic Park 1993 Spielberg, Isla Nublar at night during storm, massive Tyrannosaurus Rex looming through electrified fence torn open, jeep headlights illuminating rain, John Williams score visual panic, tropical vegetation crushed by enormous feet, DANGER HIGH VOLTAGE sign fallen, photorealistic ILM visual effects aesthetic",
    "Marvel Spider-Man New York, sunset rooftop aerial view, Manhattan skyline in golden hour, web-slinging trajectory lines visible in air, water towers and fire escapes, Empire State Building in background, heroic pose on gargoyle, Danny Elfman 2002 Sam Raimi score visual energy, photorealistic cinematic Marvel",
    "Asgard, Marvel MCU, Kenneth Branagh Thor aesthetic, golden palace with rainbow Bifrost bridge stretching across space, Norse-mythology-scale architecture, cosmos and nebula sky, Mjolnir crackling with lightning, Patrick Doyle heroic fanfare visual energy, gods-eye scale grandeur, photorealistic Marvel Studios cinematic quality",
    "Wakanda Royal Palace, Marvel Black Panther, Ryan Coogler aesthetic, vibranium crystalline architecture with Afrofuturist design, Throne Room with carved ceremonial decor, Golden City panorama through arched window, Ludwig Göransson score visual energy, advanced technology meets ancient tradition, photorealistic Marvel Studios quality",
    "Mirror Dimension, Marvel Doctor Strange, Scott Derrickson aesthetic, New York City folding and fractalizing into impossible Escher geometry, buildings rotating and multiplying in kaleidoscopic recursion, orange golden sling ring portal sparks, Michael Giacchino score visual energy, photorealistic Marvel VFX cinematic quality",
    "Ice Palace, Frozen Disney, Elsa's crystalline mountain fortress, every surface of perfect blue-white transparent ice, aurora borealis sky, snowflakes frozen mid-air, grand staircase of ice leading to open balcony, Christophe Beck Let It Go visual energy, breath-visible cold air, photorealistic Disney animation aesthetic adapted to live-action",
    "Pandora night, James Cameron Avatar, bioluminescent jungle, floating Hallelujah Mountains with waterfalls, Na'vi Tree of Souls glowing in distance, every plant emitting cyan and violet light, great leonopteryx pterodactyl silhouette in sky, James Horner score visual energy, photorealistic ILM Avatar cinematic quality",
    "Hogwarts Astronomy Tower at night, Harry Potter, grey stone battlements under brilliant star field, telescopes set up by students in robes, owlery tower visible, forbidden forest dark mass below, Black Lake reflecting castle lights, John Williams score visual wonder, photorealistic Warner Bros film aesthetic",
    "International Space Station orbital photography, Earth curve visible below, random Earth region from 408km altitude — one of: Italian peninsula boot shape, Nile delta green fan, Amazon river bends, Sahara sand dunes from above, coral reef atolls, hurricane eye, Himalayan peaks above clouds, city lights grid at night — thin atmosphere blue arc, solar panel edge in frame, curvature of Earth, weightless cosmic overview perspective",
    "H.R. Giger biomechanical corridor, deep space derelict spacecraft interior, ribbed organic walls fused with industrial metal, bioluminescent slime trails, ovomorphs eggs in pale mist on floor, motion tracker blinking green, single emergency red light beam cutting through darkness, claustrophobic tunnel perspective, Ridley Scott Alien 1979 cinematic atmosphere, photorealistic horror science fiction",
    "Xenomorph hive, Aliens 1986 James Cameron aesthetic, cavernous biomechanical chamber coated in dark organic resin secretions, hundreds of eggs in rows, faint blue-green bioluminescent glow from egg openings, cocooned forms in walls, acid burn scars on floor, atmospheric fog, hellish alien architecture, photorealistic cinematic horror",
    "Prometheus 2012 Ridley Scott, Engineer derelict spacecraft interior, vast circular orrery chamber with holographic star map projecting galaxy in centre, black stone monumental architecture, murals carved in stone showing space jockey figures, green bioluminescent light, ancient alien grandeur and cosmic dread, photorealistic IMAX cinematic quality",
    "LV-426 exomoon surface, Alien franchise atmosphere, perpetual violent electrical storm, acidic orange-brown sky with lightning, barren volcanic rock plain, crashed spacecraft silhouette on distant horizon, abandoned colony ruins, terraformer cooling towers venting steam, no breathable atmosphere, desolate hostile alien world, photorealistic science fiction",
    "Lost in Space Netflix 2018 reboot aesthetic, Jupiter 2 spacecraft landed on alien planet surface, dramatic alien sky with unusual colored moons, crystalline rock formations, Robinson family campsite with advanced equipment, Robot B9 standing sentinel nearby, alien plant life in foreground, survival camp atmosphere, photorealistic cinematic science fiction",
    "Lost in Space 2018 Netflix, dense alien forest with bioluminescent blue-white tree trunks, alien fauna in shadows, Robot B9 standing tall in dramatic light, emergency flares casting orange glow, ground covered in iridescent alien moss, survival gear scattered, atmospheric fog, photorealistic cinematic science fiction",
    "Jupiter 2 spacecraft cockpit interior, Lost in Space Netflix aesthetic, curved panoramic viewport showing alien planet atmosphere below, holographic navigation displays, advanced control panels in warm amber emergency lighting, zero-gravity handholds, survival equipment secured to walls, photorealistic hard science fiction",
    "Lost ABC TV 2004, tropical beach crash site, Oceanic Airlines Flight 815 wreckage scattered on white sand, burning fuselage section, luggage and debris in surf, tropical jungle wall behind, golden afternoon light, distant smoke column rising, survivors in background tending fires, palm trees, desperate beauty of isolation, photorealistic cinematic TV drama",
    "The Swan station, Lost ABC TV, underground DHARMA Initiative bunker, 1970s institutional aesthetic, harvest gold and avocado green appliances, pneumatic tube system, hieroglyph countdown timer on CRT monitor, long institutional corridor, DHARMA Initiative logos on walls, emergency lighting, mysterious computer terminal, photorealistic 1970s retrofuturist bunker",
    "Lost island jungle, ABC TV, dense tropical jungle interior, flickering black mechanical smoke entity moving between trees, bioluminescent vegetation, ancient carved stone wall with hieroglyphs, supernatural golden light through canopy, mysterious wheel buried in ground, photorealistic cinematic mystery drama",
    "Predator 1987 John McTiernan, dense Central American jungle canopy, oppressive humid heat haze, thermal vision perspective overlay in oranges and reds optionally visible, camouflaged Predator shimmer partially visible in canopy, mud-covered tree trunks, guerrilla camp remnants, photorealistic action film aesthetic",
    "Predator spacecraft interior, Yautja trophy room, biomechanical alien aesthetic with curved organic walls, skulls and spinal columns of various species mounted as trophies, plasma caster weapons on racks, glowing alien technology panels, ritual ceremonial lighting in copper and orange, alien civilisation war culture, photorealistic science fiction",
    "Predator 2 1990, downtown Los Angeles urban warfare night scene, rain-slicked streets, neon signs reflecting in puddles, rooftop pursuit perspective, thermal overlay optionally visible, police sirens in background, 1990s urban action cinema aesthetic, photorealistic",
    "Transformers Michael Bay aesthetic, Autobot City under Decepticon assault, enormous Cybertronian robots clashing in urban environment, explosions and vehicle transforms in motion, NEST military vehicles in foreground, dawn light breaking over destruction, photorealistic ILM visual effects cinematic quality",
    "Planet Cybertron, Transformers, metallic alien world entirely of living metal and machinery, city of Iacon with towering geometric spires, Energon energy rivers glowing blue between buildings, massive robotic citizens moving in background, binary star system in sky, epic science fiction world-building, photorealistic",
    "Sector Seven government facility, Transformers 2007, vast underground Hoover Dam bunker, giant frozen Cybertronian robot suspended in ice chamber, military personnel in hazmat suits, amber emergency lighting, classified technology racks, photorealistic US government black site aesthetic",
    "Live-action person inserted into Pixar Toy Story animated world, Mary Poppins genre-blend composite of real human with animated characters, subject standing in Andy's bedroom surrounded by animated Woody and Buzz Lightyear toys, Pixar Studio lighting on animated objects, cinematic colour grade unifying live-action and animation",
    "WALL-E 2008 Pixar aesthetic, live-action person in post-apocalyptic animated Earth landscape, towers of compressed garbage cubes, WALL-E robot companion beside subject, soft warm dusty sunset light, lonely poetry of abandoned civilisation, Mary Poppins-style human-animated hybrid composite",
    "Up 2009 Pixar, live-action person sitting on animated floating house lifted by thousands of colourful helium balloons, vast cloud panorama below, afternoon golden light, Carl's house wooden porch detail, Mary Poppins-style real human integrated into animated world, joyful impossible adventure",
    "Cars 2006 Pixar, live-action person in animated Route 66 desert town of Radiator Springs, giant talking animated cars rolling past, neon signs glowing at dusk, endless American desert highway, Mary Poppins human-cartoon hybrid world, photorealistic composite",
    "Beauty and the Beast 1991 Disney, live-action person dancing in animated enchanted ballroom, Lumiere candlestick and Cogsworth clock animated around subject, golden candlelight, enchanted rose under glass dome visible, Alan Menken score visual romance, Mary Poppins real-human-in-cartoon-world aesthetic",
    "Fantasia 1940 Disney, live-action person in animated Sorcerer's Apprentice sequence, magical brooms carrying water buckets in background, enchanted workshop flooding, Mickey Mouse's sorcerer hat nearby, Paul Dukas symphonic colour, Mary Poppins hybrid human-animation composite, classical animation cel aesthetic",
    "Frozen Disney animated world, live-action person in animated Arendelle village in winter, animated villagers, Elsa's ice palace on mountain above, magical snowflakes drifting, warm Norwegian architectural detail, Mary Poppins-style real human integrated into Disney animation world",
    "The Lion King 1994 Disney, live-action person standing on animated Pride Rock overlooking African savanna at sunrise, animated wildebeest herds below, Hans Zimmer Circle of Life visual grandeur, painted sky in gold orange purple, Mary Poppins human-cartoon world fusion, epic Disney composite",
]


def _run_lego_minifig(chat_id, uid, username, img_bytes):
    """Genera LEGO Minifig di Valeria con outfit dalla foto e location random dal pool."""
    import random
    wait = None
    try:
        wait = bot.send_message(chat_id, "🟡 <b>Creo la tua LEGO Minifig...</b>")

        # Analizza outfit dalla foto
        outfit_desc, err = analyze_scene(img_bytes, client=gemini)
        if not outfit_desc:
            try: bot.delete_message(chat_id, wait.message_id)
            except Exception: pass
            bot.send_message(chat_id,
                err or "❌ Analisi immagine fallita.",
                reply_markup=filter_keyboard("lego"))
            return

        location = random.choice(_MINIFIG_LOCATIONS)

        prompt = (
            f"Generate an ultra-detailed LEGO minifigure editorial portrait of the subject described below.\n\n"
            f"**MINIFIG SPECIFICATIONS:**\n"
            f"Classic LEGO minifigure proportions — cylindrical yellow head, rectangular torso, "
            f"short cylindrical legs, C-shaped hands, stud on top of head. "
            f"Smooth shiny plastic texture throughout. "
            f"The head must have printed facial features: oval eyes with eyelashes, defined eyebrows, "
            f"small smile, subtle makeup details.\n\n"
            f"**FACE IDENTITY — PRINTED ON MINIFIG HEAD:**\n"
            f"60-year-old Italian face characteristics printed as LEGO graphic: "
            f"short silver hair printed on head sides, silver beard printed on lower face, "
            f"octagonal glasses printed as dark frames on the face. "
            f"Same face identity as always — translated into LEGO print style.\n\n"
            f"**OUTFIT — TRANSLATED TO MINIFIG GRAPHICS:**\n"
            f"The following outfit details must be reproduced as printed LEGO torso and leg graphics:\n"
            f"{outfit_desc}\n\n"
            f"Torso print: detailed outfit graphics with exact colors from the outfit description above. "
            f"Leg print: continuation of outfit or matching accessories. "
            f"If outfit includes accessories (bags, jewelry), represent as LEGO accessory elements held in C-hands.\n\n"
            f"**LOCATION:**\n"
            f"{location}\n\n"
            f"**STYLE:**\n"
            f"Photorealistic LEGO product photography style. "
            f"The minifig is the main subject, sharply in focus, posed naturally in the location. "
            f"Location fully rendered in LEGO bricks — no real-world elements. "
            f"Bright, clean LEGO catalog lighting. "
            f"Optional: subtle depth of field with location slightly softened behind the sharp minifig.\n\n"
            f"**TECHNICAL:** 8K, editorial quality, cinematic composition.\n"
            f"WATERMARK: 'Valeria Cross AI' — champagne cursive, very small, bottom center, 90% opacity.\n"
            f"NEGATIVE: real human, realistic proportions, non-LEGO elements, dark moody lighting, "
            f"body hair, incorrect minifig proportions, non-plastic texture."
        )

        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass

        CHUNK = 3800
        chunks = [prompt[i:i+CHUNK] for i in range(0, len(prompt), CHUNK)]
        for idx, chunk in enumerate(chunks):
            header = "" if len(chunks) == 1 else f"<i>({idx+1}/{len(chunks)})</i>\n"
            bot.send_message(chat_id,
                f"{header}<code>{html.escape(chunk)}</code>",
                parse_mode="HTML")

        # Pulsanti navigazione post-prompt
        minifig_markup = types.InlineKeyboardMarkup()
        minifig_markup.row(
            types.InlineKeyboardButton("🎨 Nuovo filtro, stessa foto", callback_data="post_newfilter"),
            types.InlineKeyboardButton("🔁 Stesso filtro, nuova foto", callback_data="post_newphoto"),
        )
        minifig_markup.row(
            types.InlineKeyboardButton("🆕 Nuova foto e nuovo filtro", callback_data="post_newboth"),
        )
        bot.send_message(chat_id, "Cosa vuoi fare adesso?", reply_markup=minifig_markup)

        logger.info(f"✅ lego_minifig inviato a {username} — {location[:50]}")

    except Exception as e:
        if wait:
            try: bot.delete_message(chat_id, wait.message_id)
            except Exception: pass
        bot.send_message(chat_id,
            f"❌ <b>LEGO Minifig fallito.</b>\n<code>{html.escape(str(e))}</code>",
            parse_mode="HTML")
        logger.error(f"❌ _run_lego_minifig: {e}", exc_info=True)


def _run_lego_mosaic(chat_id, uid, username, img_bytes):
    """Converte l'immagine in mosaico LEGO con studs e la invia come PNG."""
    try:
        from PIL import Image as PILImage, ImageDraw as PILDraw

        wait = bot.send_message(chat_id, "🧱 <b>Genero il mosaico LEGO...</b>\n⏳ Attendi qualche secondo.")

        STUD = 40          # px per stud nell'immagine output
        # A3 = 420x297mm, mattoncino 1x1 = 8mm → 52x37 studs
        A3_LONG  = 52
        A3_SHORT = 37

        src = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        w, h = src.size

        # Adatta all'aspect ratio — lato lungo = 52, lato corto = 37
        if w >= h:
            cols, rows = A3_LONG, A3_SHORT
        else:
            cols, rows = A3_SHORT, A3_LONG

        src_small = src.resize((cols, rows), PILImage.LANCZOS)

        # Calcola griglia colori LEGO
        grid = []
        color_counts = {}
        lego_rgb_map = {name: (r, g, b) for name, r, g, b in LEGO_PALETTE}

        MOSAIC_TYPE = ('Plate 1×1', '3024')
        tipo_counts = {}   # {(colore, tipo, codice): qty}

        for row in range(rows):
            line = []
            for col in range(cols):
                pr, pg, pb = src_small.getpixel((col, row))
                cname = _closest_lego(pr, pg, pb)
                line.append(cname)
                color_counts[cname] = color_counts.get(cname, 0) + 1
                key = (cname, MOSAIC_TYPE[0], MOSAIC_TYPE[1])
                tipo_counts[key] = tipo_counts.get(key, 0) + 1
            grid.append(line)

        # Rendering con studs
        out_w = cols * STUD
        out_h = rows * STUD
        out = PILImage.new("RGB", (out_w, out_h), (200, 200, 200))
        draw = PILDraw.Draw(out)

        for r_idx, row in enumerate(grid):
            for c_idx, cname in enumerate(row):
                rgb = lego_rgb_map[cname]
                _draw_stud(draw, c_idx*STUD, r_idx*STUD, STUD, rgb)

        # Salva e invia
        buf = io.BytesIO()
        out.save(buf, format="PNG", optimize=True)
        buf.seek(0)

        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass

        # Pulsanti: lista mattoncini + navigazione
        lego_markup = types.InlineKeyboardMarkup()
        lego_markup.row(
            types.InlineKeyboardButton("📋 Sì, lista mattoncini", callback_data="lego_list_yes"),
            types.InlineKeyboardButton("⏭️ No grazie",           callback_data="lego_list_no"),
        )
        lego_markup.row(
            types.InlineKeyboardButton("🎨 Nuovo filtro, stessa foto", callback_data="post_newfilter"),
            types.InlineKeyboardButton("🔁 Stesso filtro, nuova foto", callback_data="post_newphoto"),
        )
        lego_markup.row(
            types.InlineKeyboardButton("🆕 Nuova foto e nuovo filtro", callback_data="post_newboth"),
        )

        bot.send_photo(chat_id, buf,
            caption=f"🧱 <b>LEGO Mosaic</b> — {cols}×{rows} studs | {sum(color_counts.values())} mattoncini totali",
            parse_mode="HTML", reply_markup=lego_markup)

        # Salva stato per eventuale lista
        lego_waiting_list[uid] = {'counts': color_counts, 'tipo_counts': tipo_counts, 'cols': cols, 'rows': rows}
        logger.info(f"✅ lego_mosaic inviato a {username} ({cols}×{rows}, {len(color_counts)} colori)")

    except Exception as e:
        logger.error(f"❌ _run_lego_mosaic: {e}", exc_info=True)
        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass
        bot.send_message(chat_id,
            f"❌ <b>LEGO Mosaic fallito.</b>\n<code>{html.escape(str(e))}</code>",
            parse_mode="HTML")

# Alias per compatibilità con i callback esistenti che chiamano _run_emoji_art
_run_emoji_art = _run_lego_mosaic


# ─── LEGO GALAXY ──────────────────────────────────────────────────────────────

def _detect_bokeh_mode(src):
    """Rileva modalità bokeh dall'analisi del centro di massa dei pixel luminosi."""
    gray = src.convert("L")
    w, h = src.size
    arr  = list(gray.getdata())
    s, n = 0, 0
    for idx, v in enumerate(arr):
        if v > 128:
            s += idx // w
            n += 1
    if n == 0:
        return 'radial'
    return 'radial' if 0.30 <= s/n/h <= 0.65 else 'vertical'


def _apply_bokeh_fast(src, mode):
    """Blur progressivo che simula bokeh F1.8 — pixel-level compositing."""
    from PIL import ImageFilter
    import math
    w, h  = src.size
    src_rgb = src.convert("RGB")
    b1 = src_rgb.filter(ImageFilter.GaussianBlur(radius=2))
    b2 = src_rgb.filter(ImageFilter.GaussianBlur(radius=4))
    b3 = src_rgb.filter(ImageFilter.GaussianBlur(radius=7))
    result = src_rgb.copy()
    for y in range(h):
        for x in range(w):
            if mode == 'radial':
                cx, cy = w/2, h/2
                dist = math.sqrt(((x-cx)/max(cx,1))**2 + ((y-cy)/max(cy,1))**2)
            else:
                dist = abs((y/max(h,1)) - 0.45) * 2.2
            dist = min(dist, 1.0)
            px = (src_rgb if dist < 0.40 else b1 if dist < 0.65 else b2 if dist < 0.85 else b3).getpixel((x,y))
            result.putpixel((x,y), px[:3])
    return result


def _is_subject_zone(x, y, cols, rows, mode):
    import math
    cx, cy = cols/2, rows/2
    if mode == 'radial':
        dist = math.sqrt(((x-cx)/max(cx,1))**2 + ((y-cy)/max(cy,1))**2)
    else:
        dist = abs((y/max(rows,1)) - 0.45) * 2.2
    return dist < 0.45


def _draw_round_bg(draw, x, y, s, rgb):
    """Cerchio piatto per sfondo bokeh — nessun bordo, solo fill."""
    shadow = tuple(max(0, c-30) for c in rgb)
    pad = int(s * 0.06)
    draw.ellipse([x+pad, y+pad, x+s-pad, y+s-pad], fill=rgb, outline=shadow)


def _draw_stud_tall(draw, x, y, size, rgb):
    """
    Stud 'alto' per zone omogenee del soggetto.
    Aggiunge un anello esterno e uno stud interno più marcato — simula un cilindro piatto.
    """
    shadow    = tuple(max(0,   c-50) for c in rgb)
    highlight = tuple(min(255, c+50) for c in rgb)
    mid_hi    = tuple(min(255, c+25) for c in rgb)

    # Anello esterno (corona)
    pad_out = int(size * 0.06)
    draw.ellipse([x+pad_out, y+pad_out, x+size-pad_out, y+size-pad_out],
                 fill=rgb, outline=shadow, width=2)
    # Highlight arco top
    draw.arc([x+pad_out, y+pad_out, x+size-pad_out, y+size-pad_out],
             start=210, end=330, fill=highlight, width=max(2, size//12))
    # Stud interno
    pad_in = int(size * 0.28)
    draw.ellipse([x+pad_in, y+pad_in, x+size-pad_in, y+size-pad_in],
                 fill=mid_hi, outline=shadow)


def _color_variation(grid, row, col, rows, cols):
    lego_rgb = {name: (r, g, b) for name, r, g, b in LEGO_PALETTE}
    cr, cg, cb = lego_rgb.get(grid[row][col], (128,128,128))
    diffs = []
    for dr, dc in [(-1,0),(1,0),(0,-1),(0,1)]:
        nr, nc = row+dr, col+dc
        if 0 <= nr < rows and 0 <= nc < cols:
            r2,g2,b2 = lego_rgb.get(grid[nr][nc], (128,128,128))
            diffs.append(abs(cr-r2)+abs(cg-g2)+abs(cb-b2))
    return sum(diffs)/len(diffs) if diffs else 0


def _run_lego_galaxy(chat_id, uid, username, img_bytes):
    """
    LEGO Galaxy: sfondo bokeh → round piatti, soggetto → stud standard/tall.
    Stesso formato flat del Mosaic ma con tre elementi visivi distinti.
    """
    try:
        from PIL import Image as PILImage, ImageDraw as PILDraw

        wait = bot.send_message(chat_id,
            "🌌 <b>Genero il LEGO Galaxy...</b>\n⏳ Attendi qualche secondo.")

        STUD    = 40
        A3_LONG  = 52
        A3_SHORT = 37

        src = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        w, h = src.size

        bokeh_mode = _detect_bokeh_mode(src)
        logger.info(f"🌌 lego_galaxy: bokeh_mode={bokeh_mode}")

        if w >= h:
            cols, rows = A3_LONG, A3_SHORT
        else:
            cols, rows = A3_SHORT, A3_LONG

        src_bokeh      = _apply_bokeh_fast(src, bokeh_mode)
        src_small      = src_bokeh.resize((cols, rows), PILImage.LANCZOS)
        src_orig_small = src.resize((cols, rows), PILImage.LANCZOS)

        lego_rgb_map = {name: (r, g, b) for name, r, g, b in LEGO_PALETTE}
        color_counts = {}

        # Codici BrickLink per i tre tipi
        BRICK_TYPES = {
            'round_bg':  ('Plate, Round 1×1', '4073'),
            'standard':  ('Plate 1×1',        '3024'),
            'tall':      ('Brick, Round 1×1', '3062b'),
        }
        tipo_counts = {}   # {(colore, tipo, codice): qty}

        grid = []
        grid_types = []   # parallela a grid — tipo elemento per cella
        for row in range(rows):
            line, tline = [], []
            for col in range(cols):
                if _is_subject_zone(col, row, cols, rows, bokeh_mode):
                    pr, pg, pb = src_orig_small.getpixel((col, row))
                else:
                    pr, pg, pb = src_small.getpixel((col, row))
                line.append(_closest_lego(pr, pg, pb))
                tline.append(None)  # riempito nel rendering
            grid.append(line)
            grid_types.append(tline)

        out_w = cols * STUD
        out_h = rows * STUD
        out   = PILImage.new("RGB", (out_w, out_h), (20, 20, 20))
        draw  = PILDraw.Draw(out)

        for row in range(rows):
            for col in range(cols):
                cname   = grid[row][col]
                rgb     = lego_rgb_map[cname]
                px, py  = col*STUD, row*STUD
                is_subj = _is_subject_zone(col, row, cols, rows, bokeh_mode)
                color_counts[cname] = color_counts.get(cname, 0) + 1

                if not is_subj:
                    _draw_round_bg(draw, px, py, STUD, rgb)
                    tkey = 'round_bg'
                else:
                    var = _color_variation(grid, row, col, rows, cols)
                    if var <= 25:
                        _draw_stud_tall(draw, px, py, STUD, rgb)
                        tkey = 'tall'
                    else:
                        _draw_stud(draw, px, py, STUD, rgb)
                        tkey = 'standard'

                tname, tcode = BRICK_TYPES[tkey]
                key = (cname, tname, tcode)
                tipo_counts[key] = tipo_counts.get(key, 0) + 1

        buf = io.BytesIO()
        out.save(buf, format="PNG", optimize=True)
        buf.seek(0)

        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass

        lego_markup = types.InlineKeyboardMarkup()
        lego_markup.row(
            types.InlineKeyboardButton("📋 Sì, lista mattoncini", callback_data="lego_list_yes"),
            types.InlineKeyboardButton("⏭️ No grazie",           callback_data="lego_list_no"),
        )
        lego_markup.row(
            types.InlineKeyboardButton("🎨 Nuovo filtro, stessa foto", callback_data="post_newfilter"),
            types.InlineKeyboardButton("🔁 Stesso filtro, nuova foto", callback_data="post_newphoto"),
        )
        lego_markup.row(
            types.InlineKeyboardButton("🆕 Nuova foto e nuovo filtro", callback_data="post_newboth"),
        )

        total = sum(color_counts.values())
        bot.send_photo(chat_id, buf,
            caption=(f"🌌 <b>LEGO Galaxy</b> — {cols}×{rows} studs | "
                     f"{total} elementi | bokeh: {bokeh_mode}"),
            parse_mode="HTML", reply_markup=lego_markup)

        lego_waiting_list[uid] = {'counts': color_counts, 'tipo_counts': tipo_counts, 'cols': cols, 'rows': rows}
        logger.info(f"✅ lego_galaxy inviato a {username} ({cols}×{rows}, bokeh={bokeh_mode})")

    except Exception as e:
        logger.error(f"❌ _run_lego_galaxy: {e}", exc_info=True)
        try: bot.delete_message(chat_id, wait.message_id)
        except Exception: pass
        bot.send_message(chat_id,
            f"❌ <b>LEGO Galaxy fallito.</b>\n<code>{html.escape(str(e))}</code>",
            parse_mode="HTML")

def _run_generation(chat_id, uid, username, data):
    fkey = data['filter_key']
    # Per artistic_style: usa artist_key da data o da user_artist
    if fkey == "artistic_style":
        artist_key = data.get('artist_key') or user_artist.get(uid)
        if artist_key and artist_key in ARTISTIC_STYLE_PROMPTS:
            filter_prompt = ARTISTIC_STYLE_PROMPTS[artist_key]
            fname = f"🎨 {artist_key}"
        else:
            # fallback casuale
            filter_prompt, style_label = build_artistic_style_prompt()
            fname = f"🎨 {style_label.split(' — ')[0]}"
        pending_artistic_style.pop(uid, None)
    else:
        filter_prompt = resolve_prompt(fkey)
        fname = FILTERS[fkey]["label"]
    img_bytes = data['img']

    t_start = time.time()
    logger.info(f"   🎨 Inizio generazione | {username} | filtro: {fkey}")

    # --- ANALISI OUTFIT REALE (fino a 2 tentativi) ---
    wait_analysis = bot.send_message(chat_id, "🔍 <b>Analizzo l'outfit...</b>")
    outfit_desc, err = analyze_scene(img_bytes, gemini)
    try: bot.delete_message(chat_id, wait_analysis.message_id)
    except Exception: pass

    if not outfit_desc:
        bot.send_message(chat_id,
            f"❌ <b>Analisi outfit fallita dopo 2 tentativi.</b>\n"
            f"Nessun prompt generato.\n\n{err or ''}"
        )
        logger.error(f"   ❌ _run_generation: analyze_scene fallita per {username} | filtro: {fkey}")
        return

    # Wrapper editoriale per ridurre falsi positivi IMAGE_SAFETY
    EDITORIAL_WRAPPER = (
        "This is a professional editorial post-production request for a high-fashion photography project. "
        "This is a legitimate creative and commercial photography workflow. "
        "Apply the following filter: "
    )

    try:
        full_prompt = (
            f"SUBJECT REFERENCE — LOCKED — DO NOT ALTER:\n"
            f"The following description of the subject was extracted from the photo. "
            f"Reproduce the subject, outfit, accessories and setting exactly as described. "
            f"DO NOT alter, replace, invent or omit any detail.\n\n"
            f"{outfit_desc}\n\n"
            f"{filter_prompt}"
        )
        last_prompt[uid] = full_prompt

        elapsed = round(time.time() - t_start, 1)

        # Invia prompt Flow-ready
        post_markup = types.InlineKeyboardMarkup()
        post_markup.row(
            types.InlineKeyboardButton("🎨 Nuovo filtro, stessa foto", callback_data="post_newfilter"),
            types.InlineKeyboardButton("🔁 Stesso filtro, nuova foto", callback_data="post_newphoto")
        )
        post_markup.row(
            types.InlineKeyboardButton("🆕 Nuova foto e nuovo filtro", callback_data="post_newboth")
        )

        bot.send_message(chat_id, f"⚡ <b>Prompt pronto.</b> Usa Flow per generare l'immagine.")
        chunks = [full_prompt[i:i+3800] for i in range(0, len(full_prompt), 3800)]
        for idx, chunk in enumerate(chunks):
            header = "" if len(chunks) == 1 else f"<i>({idx+1}/{len(chunks)})</i>\n"
            if idx == len(chunks) - 1:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>",
                    reply_markup=post_markup)
            else:
                bot.send_message(chat_id, f"{header}<code>{html.escape(chunk)}</code>")

        bot.send_message(chat_id,
            "⚠️ <b>Ricorda:</b> per risultati fedeli all'originale, carica l'immagine di riferimento "
            "su Flow insieme a questo prompt.",
            parse_mode="HTML")



        logger.info(f"   ✅ Prompt inviato a {username} in {elapsed}s")
        return

    except Exception as e:
        elapsed = round(time.time() - t_start, 1)
        logger.error(f"   ❌ Errore generazione prompt ({elapsed}s): {e}", exc_info=True)
        bot.send_message(chat_id, f"❌ Errore interno:\n<code>{html.escape(str(e))}</code>")

# --- MAIN ---

# --- /shared ---
@bot.message_handler(commands=['shared'])
def cmd_shared(m):
    bot.send_message(m.chat.id,
        f"📦 <b>C_shared100.py</b> v{SHARED_VERSION} — {SHARED_DATE}"
    )


@bot.callback_query_handler(func=lambda c: c.data in ("lego_list_yes", "lego_list_no"))
def cb_lego_list(call):
    uid = call.from_user.id
    if not is_allowed(uid): return
    try: bot.answer_callback_query(call.id)
    except Exception: pass
    try: bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception: pass

    if call.data == "lego_list_no":
        return

    data = lego_waiting_list.pop(uid, None)
    if not data:
        bot.send_message(call.message.chat.id, "⚠️ Lista non disponibile. Genera di nuovo il mosaico.")
        return

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        cols_n, rows_n = data['cols'], data['rows']
        tipo_counts = data.get('tipo_counts', {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista Mattoncini"

        # Intestazione
        ws['A1'] = f"LEGO Mosaic — {cols_n}×{rows_n} studs"
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:E1')

        headers = ["Colore LEGO", "Tipo elemento", "Codice BrickLink", "Quantità", "Link BrickLink"]
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style='thin', color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Colori LEGO per l'evidenziazione
        LEGO_HEX = {
            "White": "F2F3F2", "Light Bluish Gray": "ABABAC", "Medium Stone Grey": "A3A2A5",
            "Dark Bluish Gray": "6C6E68", "Dark Stone Grey": "4A4948", "Black": "1B2A34",
            "Yellow": "F7D117", "Bright Yellow": "FFCD00", "Warm Gold": "AA7F2E",
            "Tan": "DEC69C", "Light Nougat": "F9C198", "Nougat": "CC8E57",
            "Medium Nougat": "AA7D55", "Dark Nougat": "AA550A", "Reddish Brown": "69402A",
            "Dark Brown": "351F00", "Red": "B3000A", "Bright Red": "C80008",
            "Dark Red": "720E0F", "Orange": "D56316", "Bright Orange": "FF7E14",
            "Dark Orange": "9D4F0C", "Coral": "FF6D58", "Salmon": "FF9077",
            "Light Pink": "FFADB9", "Pink": "FF698F", "Magenta": "901F76",
            "Dark Pink": "C90046", "Bright Green": "00B900", "Green": "00924A",
            "Dark Green": "00642E", "Lime Green": "7DBD00", "Medium Green": "ACCD98",
            "Olive Green": "4B5420", "Sand Green": "78988A", "Blue": "0055BF",
            "Bright Blue": "0072CE", "Dark Blue": "0020A0", "Medium Blue": "61AFFF",
            "Sky Blue": "87CDE2", "Dark Azure": "0083CB", "Medium Azure": "00A8DE",
            "Sand Blue": "5F7C94", "Purple": "510079", "Medium Lilac": "342B75",
            "Lavender": "A0968C", "Medium Lavender": "A97FC8", "Sand Purple": "877880",
        }

        # Dati ordinati per quantità decrescente
        sorted_items = sorted(tipo_counts.items(), key=lambda x: x[1], reverse=True)

        alt_fill_a = PatternFill("solid", fgColor="F8F9FA")
        alt_fill_b = PatternFill("solid", fgColor="FFFFFF")

        for r_idx, ((color, tipo, codice), qty) in enumerate(sorted_items, 3):
            link = f"https://www.bricklink.com/v2/catalog/catalogitem.page?P={codice}"
            row_fill = alt_fill_a if r_idx % 2 == 0 else alt_fill_b

            # Cella colore con sfondo LEGO
            hex_c = LEGO_HEX.get(color, "EEEEEE")
            # Testo chiaro/scuro in base alla luminosità
            r2,g2,b2 = int(hex_c[0:2],16), int(hex_c[2:4],16), int(hex_c[4:6],16)
            lum = 0.299*r2 + 0.587*g2 + 0.114*b2
            txt_color = "000000" if lum > 128 else "FFFFFF"

            c1 = ws.cell(row=r_idx, column=1, value=color)
            c1.fill = PatternFill("solid", fgColor=hex_c)
            c1.font = Font(bold=True, color=txt_color)
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c1.border = border

            for c_idx, val in enumerate([tipo, codice, qty], 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Link BrickLink come hyperlink
            link_cell = ws.cell(row=r_idx, column=5, value="🔗 BrickLink")
            link_cell.hyperlink = link
            link_cell.font = Font(color="0055BF", underline="single")
            link_cell.fill = row_fill
            link_cell.alignment = Alignment(horizontal='center', vertical='center')
            link_cell.border = border

        # Riga totale
        total = sum(tipo_counts.values())
        n_colors = len(set(k[0] for k in tipo_counts))
        n_tipi   = len(set(k[1] for k in tipo_counts))
        tot_row  = len(sorted_items) + 3
        ws.cell(row=tot_row, column=1, value="TOTALE").font = Font(bold=True)
        ws.cell(row=tot_row, column=4, value=total).font = Font(bold=True)
        ws.cell(row=tot_row, column=1).fill = PatternFill("solid", fgColor="1A1A2E")
        ws.cell(row=tot_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=tot_row, column=4).fill = PatternFill("solid", fgColor="1A1A2E")
        ws.cell(row=tot_row, column=4).font = Font(bold=True, color="FFFFFF")
        for ci in range(1, 6):
            ws.cell(row=tot_row, column=ci).border = border

        # Info aggiuntiva
        ws.cell(row=tot_row+1, column=1, value=f"{n_colors} colori · {n_tipi} tipi elemento")
        ws.cell(row=tot_row+1, column=1).font = Font(italic=True, color="666666")
        ws.merge_cells(f'A{tot_row+1}:E{tot_row+1}')

        # Larghezze colonne
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 16

        # Freeze header
        ws.freeze_panes = "A3"

        # Salva e invia
        buf_xl = io.BytesIO()
        wb.save(buf_xl)
        buf_xl.seek(0)

        bot.send_document(call.message.chat.id, buf_xl,
            visible_file_name=f"LEGO_lista_{cols_n}x{rows_n}.xlsx",
            caption=f"📊 <b>Lista mattoncini</b> — {total} pz in {n_colors} colori, {n_tipi} tipi",
            parse_mode="HTML")
        logger.info(f"📊 Lista LEGO Excel inviata a {call.from_user.username} ({total} pz)")

    except Exception as e:
        logger.error(f"❌ cb_lego_list Excel: {e}", exc_info=True)
        bot.send_message(call.message.chat.id,
            f"❌ Errore generazione Excel:\n<code>{html.escape(str(e))}</code>",
            parse_mode="HTML")


if __name__ == "__main__":
    import time
    logger.info(f"🎨 Avvio Filtro v{VERSION}")
    server.start()
    if not gemini.available:
        logger.warning("⚠️ GOOGLE_API_KEY non configurata.")
    while True:
        try:
            bot.infinity_polling(timeout=30, long_polling_timeout=25)
        except Exception as e:
            err = str(e)
            if "409" in err or "Conflict" in err:
                logger.warning("⚠️ 409 Conflict — altra istanza attiva. Attendo 15s e riprovo...")
                time.sleep(15)
            else:
                logger.error(f"❌ Polling error: {e}")
                time.sleep(5)
